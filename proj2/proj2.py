# Imported libraries required for programme (math, pandas, numpy, os, shutil) and cleared screen using cls command
import streamlit as st
import numpy as np
from calendar import c
import math
import os
import pandas as pd
import numpy as np
import shutil
import pandas.io.formats.excel
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border,Side
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
import concurrent.futures
from os import listdir
from os.path import isfile, join
from datetime import datetime
start_time = datetime.now()
import os
import pandas as pd
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb
import streamlit as st
# Change Directory here
os.chdir(r'C:\Users\pc\Documents\GitHub\2001CE45_2022\proj2')

# Function to get BytesIO
def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    writer.save()
    processed_data = output.getvalue()
    return processed_data

# Function defination
# @st.experimental_memo(suppress_st_warning=True)
def proj_octant_gui():

	
	# Title of Webpage
	st.title("Processed Output File Generator")

	# Menu for Sidebar
	menu = ["Single/Multiple File Upload","Bulk File Upload"]
	choice = st.sidebar.selectbox("Menu", menu)

	# Single or Multiple File Upload
	if choice == "Single/Multiple File Upload":
		st.subheader("Single/Multiple File Upload")
		uploaded_files = st.file_uploader("Choose a Excel file/files", accept_multiple_files=True, type = ['csv','xlsx'])
		for uploaded_file in uploaded_files:
			bytes_data = uploaded_file.read()
			# st.write("filename:", uploaded_file.name)
			# st.write(bytes_data)

		# Taking mod input from user
		mod = st.number_input('Please Enter a Mod input', value = 5000)
		
		st.write("Please Click on ***Compute*** for Computing Output:")
		# Compute Checkbox
		if st.checkbox("Compute"):
			# Iterating through uploaded files
			for uploaded_file in uploaded_files:
				# Running tut 7 code for uploaded files and creating output file
				# st.write(uploaded_file)
				# Creating dictionary for octant name id mapping
				octant_name_id_mapping = {"1":"Internal outward interaction",
				"-1":"External outward interaction",
				"2":"External Ejection",
				"-2":"Internal Ejection",
				"3":"External inward interaction",
				"-3":"Internal inward interaction",
				"4":"Internal sweep",
				"-4":"External sweep"
				}
				# Creating dataframe from excel
				df = pd.read_excel(uploaded_file)
				size_a = 200
				# size_a=df['U'].size
				# Calculating no. of iterations for mod ranges using ceil function of Math library
				try:
					iter = math.ceil(size_a/ mod)
					total_count=0
				# ZeroDivisionError for division by zero
				except ZeroDivisionError:
					print("Mod value cannot be zero! Please Change Mod value!!!")
					exit()
				# Created size variable and u_mean, v_mean, w_mean variables and stored means of individual columns in that
				
				u_mean=(df['U']).mean()
				v_mean=(df['V']).mean()
				w_mean=(df['W']).mean()

				# Try Block if Output file not found then to except
				try:
					#Created new blank columns for average velocities
					df["U avg"]= np.nan
					df["V avg"]= np.nan
					df["W avg"]= np.nan

					# Stored respective means at their position and column
					# .loc[] is used to access a group of rows and columns by labels
					df.loc[0,'U avg']=round(u_mean,3)
					df.loc[0,'V avg']=round(v_mean,3)
					df.loc[0,'W avg']=round(w_mean,3)

					# Created empty columns for U' = U- U avg
					# np.nan fills the column with NaN values
					df["U' = U - U avg"]= np.nan
					# Filled empty columns with required values of difference between U and U avg
					for i in range(0, size_a):
						df.loc[i,"U' = U - U avg"]= round(df.loc[i,'U']-u_mean,3)
					
					# Created empty columns for V' = V- V avg
					df["V' = V - V avg"]= np.nan
					# Filled empty columns with required values of difference between V and V avg
					for i in range(0, size_a):
						df.loc[i,"V' = V - V avg"]= round(df.loc[i,'V']-v_mean,3)

					# Created empty columns for W' = W- W avg
					df["W' = W - W avg"]= np.nan
					# Filled empty columns with required values of difference between W and W avg
					for i in range(0, size_a):
						df.loc[i,"W' = W - W avg"]= round(df.loc[i,'W']-w_mean,3)

					# Created an empty column for Octant Value storage
					df["Octant_value"]=np.nan
					
					# Logic for storing Octant Values
					# '''Iterated through all rows and compared different permutations and assigned Octant Value'''
					# Running Loop for every row
					for i in range(0, size_a):
						# Comparison for selecting octant value
						if df.loc[i,"U' = U - U avg"]>=0 and df.loc[i,"V' = V - V avg"]>=0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=1
							
							else:
								df.loc[i,"Octant_value"]=-1
						
						elif df.loc[i,"U' = U - U avg"]<0 and df.loc[i,"V' = V - V avg"]>=0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=2
							
							else:
								df.loc[i,"Octant_value"]=-2
						
						elif df.loc[i,"U' = U - U avg"]<0 and df.loc[i,"V' = V - V avg"]<0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=3
							
							else:
								df.loc[i,"Octant_value"]=-3
						
						else:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=4
							
							else:
								df.loc[i,"Octant_value"]=-4
					# st.write(df)
					# Typecasting Octant values from float to int
					df["Octant_value"]=df["Octant_value"].astype('Int64')
					
					# Created column
					df[""]=np.nan 
					
					# Just Created Good interphase for Excel file
					df.loc[0,"Octant ID"]="Overall Octant Count"

					# Initialized count variables with zero   
					count_pos1=0
					count_neg1=0
					count_pos2=0
					count_neg2=0
					count_pos3=0
					count_neg3=0
					count_pos4=0
					count_neg4=0
						
					# Logic for counting octant values
					# Iterated through Octant Value column and compared with octant value and increased count by one        
					
					for i in range(0,size_a):
						# print(type(df.loc[i,"Octant_value"]))
						if df.loc[i,"Octant_value"]==1:
							count_pos1 +=1
						elif df.loc[i,"Octant_value"]==-1:
							count_neg1 +=1
						elif df.loc[i,"Octant_value"]==2:
							count_pos2 +=1
						elif df.loc[i,"Octant_value"]==-2:
							count_neg2 +=1
						elif df.loc[i,"Octant_value"]==3:
							count_pos3 +=1
						elif df.loc[i,"Octant_value"]==-3:
							count_neg3 +=1
						elif df.loc[i,"Octant_value"]==4:
							count_pos4 +=1
						elif df.loc[i, "Octant_value"]==-4:
							count_neg4 +=1
				
					# Written values of counts at specified positions
					df.loc[0,"1"]=count_pos1
					df.loc[0,"-1"]=count_neg1
					df.loc[0,"2"]=count_pos2
					df.loc[0,"-2"]=count_neg2
					df.loc[0,"3"]=count_pos3
					df.loc[0,"-3"]=count_neg3        
					df.loc[0,"4"]=count_pos4   
					df.loc[0,"-4"]=count_neg4
					
					# Dictionary for mapping index to octant values
					util={
						1:1,
						2:-1,
						3:2,
						4:-2,
						5:3,
						6:-3,
						7:4,
						8:-4
					}
					# Logic: Creating a list for count values and creaating a copy of it. Sorting the copy of list.
					# Creating Dictionary of copy list and storing values as keys, indices as values
					ls = [count_pos1, count_neg1, count_pos2, count_neg2, count_pos3, count_neg3, count_pos4, count_neg4]
					ls_ = ls.copy()
					ls_.sort(reverse=True)
					ls_dict = {k: v for v, k in enumerate(ls_)}
					
					# Then Storing Ranks at specified positions
					for m in range(8):
						df.loc[0,f'Rank Octant {util[m+1]}']=ls_dict[ls[m]]+1
						if ls_dict[ls[m]]==0:
							store=m
					# Creating Structure and storing Rank 1's Octant ID and its name
					df.loc[0,"Rank 1 Octant ID"] = util[store+1]
					df.loc[0, "Rank 1 Octant Name"] = octant_name_id_mapping[str(util[store+1])]
					
					s = "Mod "+ str(mod)
					df.loc[0,""]= s

					
					# Creating a list for storing mod Ranks
					ranklist=[]
					# Iterating through each range and count the octant value of each ID
					for i in range(2, iter+2):
						# Writing range in file
						stg=""
						if i==2:
							x = min(mod*(i-1)-1,size_a)
							stg = "0000"+"-"+str(x)
						elif i== iter+1:
							stg= str(mod*(i-2))+"-"+str(size_a-1)
						else:
							stg = str(mod *(i-2))+"-"+str(mod*(i-1)-1)
						df.loc[i-1,"Octant ID"]= stg

						# Initializing mod counts as zeroes
						mod_cnt_pos_one=0
						mod_cnt_neg_one=0
						mod_cnt_pos_two=0
						mod_cnt_neg_two=0
						mod_cnt_pos_three=0
						mod_cnt_neg_three=0
						mod_cnt_pos_four=0
						mod_cnt_neg_four=0

						# Handling current range for last call
						current_range= min(size_a, (i-1)*mod)

						# Iterating over range and incrementing count for respective mod count
						for j in range((i-2)*mod, current_range):
							if df.loc[j, "Octant_value"]==1:
								mod_cnt_pos_one+=1
							elif df.loc[j, "Octant_value"]==-1:
								mod_cnt_neg_one+=1
							elif df.loc[j, "Octant_value"]==2:
								mod_cnt_pos_two+=1
							elif df.loc[j, "Octant_value"]==-2:
								mod_cnt_neg_two+=1
							elif df.loc[j, "Octant_value"]==3:
								mod_cnt_pos_three+=1
							elif df.loc[j, "Octant_value"]==-3:
								mod_cnt_neg_three+=1
							elif df.loc[j, "Octant_value"]==4:
								mod_cnt_pos_four+=1
							elif df.loc[j, "Octant_value"]==-4:
								mod_cnt_neg_four+=1
						
						# writing mod counts at specified locations
						df.loc[i-1, "1"]=mod_cnt_pos_one
						df.loc[i-1, "-1"]=mod_cnt_neg_one
						df.loc[i-1, "2"]=mod_cnt_pos_two
						df.loc[i-1, "-2"]=mod_cnt_neg_two
						df.loc[i-1, "3"]=mod_cnt_pos_three
						df.loc[i-1, "-3"]=mod_cnt_neg_three
						df.loc[i-1, "4"]=mod_cnt_pos_four
						df.loc[i-1, "-4"]=mod_cnt_neg_four
						
						# Logic: Creating a list for count values and creaating a copy of it. Sorting the copy of list.
						# Creating Dictionary of copy list and storing values as keys, indices as values
						ls = [mod_cnt_pos_one, mod_cnt_neg_one, mod_cnt_pos_two, mod_cnt_neg_two, mod_cnt_pos_three, mod_cnt_neg_three, mod_cnt_pos_four, mod_cnt_neg_four]
						ls_ = ls.copy()
						ls_.sort(reverse=True)
						ls_dict = {k: v for v, k in enumerate(ls_)}
						# Then Storing Ranks at specified positions
						for m in range(8):
							df.loc[i-1,f'Rank Octant {util[m+1]}']=ls_dict[ls[m]]+1
							if ls_dict[ls[m]]==0:
								store=m
						# Creating structure and storing Rank 1's Octant ID
						df.loc[i-1,"Rank 1 Octant ID"] = util[store+1]
						# Appending current Rank's octant to ranklist
						ranklist.append(util[store+1])
						# Storing current Octant's name
						df.loc[i-1, "Rank 1 Octant Name"] = octant_name_id_mapping[str(util[store+1])]
					
					# Creating structure 
					df.loc[iter+2,"Rank Octant 4"]="Octant ID"
					df.loc[iter+2,"Rank Octant -4"]="Octant Name"
					df.loc[iter+2,"Rank 1 Octant ID"]="Count of Rank 1 Mod Values"
					# Iterating over loop and storing mod counts of each octant
					for n in range(iter+3, iter+11):
						df.loc[n,"Rank Octant 4"]= (util[n-iter-2])
						df.loc[n,"Rank Octant -4"]=octant_name_id_mapping[str(util[n-iter-2])]
						df.loc[n,"Rank 1 Octant ID"]= ranklist.count(util[n-iter-2])
					
					

				# Except block if File not found or incorrect directory is opened
				except FileNotFoundError:
					st.write("Oops! Error Occured! File Not Found!")
					st.write("Please Check File present in correct directory with correct name")
					exit()

				#Creating Structure for Overall Transition count Table
				df["_"]=np.nan
				df["Overall Transition Count"]=np.nan
				df.loc[1,"Overall Transition Count"]="Octant"
				df.loc[0,"1_"]="To"
				df.loc[1,"_"]="From"
				df.loc[2,"Overall Transition Count"]="+1"
				df.loc[3,"Overall Transition Count"]="-1"
				df.loc[4,"Overall Transition Count"]="+2"
				df.loc[5,"Overall Transition Count"]="-2"
				df.loc[6,"Overall Transition Count"]="+3"
				df.loc[7,"Overall Transition Count"]="-3"
				df.loc[8,"Overall Transition Count"]="+4"
				df.loc[9,"Overall Transition Count"]="-4"

				df.loc[1,"1_"]="+1"
				df.loc[1,"-1_"]="-1"
				df.loc[1,"2_"]="+2"
				df.loc[1,"-2_"]="-2"
				df.loc[1,"3_"]="+3"
				df.loc[1,"-3_"]="-3"
				df.loc[1,"4_"]="+4"
				df.loc[1,"-4_"]="-4"

				# Initializing every cell in table as 0
				for i in range(8):
					df.loc[2+i,"1_"]=0
					df.loc[2+i,"-1_"]=0
					df.loc[2+i,"2_"]=0
					df.loc[2+i,"-2_"]=0
					df.loc[2+i,"3_"]=0
					df.loc[2+i,"-3_"]=0
					df.loc[2+i,"4_"]=0
					df.loc[2+i,"-4_"]=0


				# Counting and filling the table of overall transition count
				for i in range(0,size_a-1):
					x= df.loc[i,"Octant_value"]
					y=df.loc[i+1,"Octant_value"]
					z=0
					if x>0:
						z= x*2 -1
					elif x<0:
						z= -2 * x
					
					df.loc[1+z,f'{y}_']=df.loc[1+z, f'{y}_']+1

				# Counting the mod ranges using this for loop
				for i in range(iter):
					# Creating structures for mod ranges count
					df.loc[iter+19+13*i-iter-6, "Overall Transition Count"]="Mod Transition Count"
					df.loc[iter+21+13*i-iter-6,"Overall Transition Count"]="Octant"
					df.loc[iter+22+13*i-iter-6,"_"]="From"
					df.loc[iter+20+13*i-iter-6,"1_"]="To"
					df.loc[iter+22+13*i-iter-6,"Overall Transition Count"]="+1"
					df.loc[iter+23+13*i-iter-6,"Overall Transition Count"]="-1"
					df.loc[iter+24+13*i-iter-6,"Overall Transition Count"]="+2"
					df.loc[iter+25+13*i-iter-6,"Overall Transition Count"]="-2"
					df.loc[iter+26+13*i-iter-6,"Overall Transition Count"]="+3"
					df.loc[iter+27+13*i-iter-6,"Overall Transition Count"]="-3"
					df.loc[iter+28+13*i-iter-6,"Overall Transition Count"]="+4"
					df.loc[iter+29+13*i-iter-6,"Overall Transition Count"]="-4"

					df.loc[iter+21+13*i-iter-6,"1_"]="+1"
					df.loc[iter+21+13*i-iter-6,"-1_"]="-1"
					df.loc[iter+21+13*i-iter-6,"2_"]="+2"
					df.loc[iter+21+13*i-iter-6,"-2_"]="-2"
					df.loc[iter+21+13*i-iter-6,"3_"]="+3"
					df.loc[iter+21+13*i-iter-6,"-3_"]="-3"
					df.loc[iter+21+13*i-iter-6,"4_"]="+4"
					df.loc[iter+21+13*i-iter-6,"-4_"]="-4"

					# Initializing every cell of table to zero
					for j in range(8):
						df.loc[iter+22+j+13*i-iter-6,"1_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-1_"]=0
						df.loc[iter+22+j+13*i-iter-6,"2_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-2_"]=0
						df.loc[iter+22+j+13*i-iter-6,"3_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-3_"]=0
						df.loc[iter+22+j+13*i-iter-6,"4_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-4_"]=0

					# Initializing start and end variables
					start = i * mod
					end = ((i+1) * mod) if ((i+1) * mod) < size_a else size_a-1
					# Creating string for mod ranges and storing at correct location
					if i==0:
						x = min(mod*(i+1),size_a)
						s = "0000"+"-"+str(x-1)
					elif i== iter-1:
						s= str(start)+"-"+str(size_a-1)
					else:
						s = str(start)+"-"+str(end-1)
					df.loc[iter+20+13*i-iter-6,"Overall Transition Count"]= s
					
					#Storing count of transition in cell of table after iterating through range 
					for k in range(start,end):
						x= df.loc[k,"Octant_value"]
						y= df.loc[k+1,"Octant_value"]
						z=0
						if x>0:
							z= x*2 -1
						elif x<0:
							z= -2 * x
						
						df.loc[iter+21+13*i+z-iter-6,f'{y}_']=df.loc[iter+21+z+13*i-iter-6, f'{y}_']+1
				
				df["Octant"]=np.nan
				df.loc[0, "Octant"]="+1"
				df.loc[1, "Octant"]="-1"
				df.loc[2, "Octant"]="+2"
				df.loc[3, "Octant"]="-2"
				df.loc[4, "Octant"]="+3"
				df.loc[5, "Octant"]="-3"
				df.loc[6, "Octant"]="+4"
				df.loc[7, "Octant"]="-4"

				# Created columns Longest subsequence length and count
				df["Longest Subsequence Length"]=np.nan
				df["Count"]=np.nan

				# Initialized variables for max length of subsequence of each octant
				max_length_p1=0
				max_length_n1=0
				max_length_p2=0
				max_length_n2=0
				max_length_p3=0
				max_length_n3=0
				max_length_p4=0
				max_length_n4=0

				# Initialized Count variables for each octant value
				count_p1 = 0
				count_n1 = 0
				count_p2 = 0
				count_n2 = 0
				count_p3 = 0
				count_n3 = 0
				count_p4 = 0
				count_n4 = 0
				# Initialized time range variables of octant as list to store j
				time_rangesp1 = []
				time_rangesn1 = []
				time_rangesp2 = []
				time_rangesn2 = []
				time_rangesp3 = []
				time_rangesn3 = []
				time_rangesp4 = []
				time_rangesn4 = []


				# Iterated through overall range of octant columns
				for j in range(size_a):
					# Initialized curr_length variable as 0 for keep track of length of current octant value subsequence
					curr_length = 0
					# Variable x for storing Octant Value
					x = df.loc[j,"Octant_value"]
					# Current index as idx
					idx = j

					# Comparing current octant value with each 8 octant values and incrementing corresponding 
					# max subsequence length variable and count of that max subsequence
					# For 1
					if x==1:
						# Increasing current length by 1
						curr_length +=1
						# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
						if curr_length>max_length_p1:
							count_p1=1
							# Clearing time range if new max length found and appending that index to index range
							time_rangesp1.clear()
							time_rangesp1.append([df.loc[idx, "T"], df.loc[idx , "T"]])
						# If current length is equal to the maximum length then increase the count of maximum subsequence length
						elif curr_length==max_length_p1:
							count_p1+=1
							# Appending the list for from time to time as list to time range for current octant
							time_rangesp1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						# storing maximum subsequence length by comparing it with current length
						max_length_p1 = max(max_length_p1, curr_length)
					
					# For -1
					if x==-1:
						curr_length +=1
						if curr_length>max_length_n1:
							count_n1=1
							time_rangesn1.clear()
							time_rangesn1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n1:
							count_n1+=1
							time_rangesn1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n1 = max(max_length_n1, curr_length)

					# For 2
					if x==2:
						curr_length +=1
						if curr_length>max_length_p2:
							count_p2=1
							time_rangesp2.clear()
							time_rangesp2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p2:
							count_p2+=1
							time_rangesp2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p2 = max(max_length_p2, curr_length)
					# For -2
					if x==-2:
						curr_length +=1
						if curr_length>max_length_n2:
							count_n2=1
							time_rangesn2.clear()
							time_rangesn2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n2:
							count_n2+=1
							time_rangesn2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n2 = max(max_length_n2, curr_length)
					# For 3
					if x==3:
						curr_length +=1
						if curr_length>max_length_p3:
							count_p3=1
							time_rangesp3.clear()
							time_rangesp3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p3:
							count_p3+=1
							time_rangesp3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p3 = max(max_length_p3, curr_length)
					# For -3
					if x==-3:
						curr_length +=1
						if curr_length>max_length_n3:
							count_n3=1
							time_rangesn3.clear()
							time_rangesn3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n3:
							count_n3+=1
							time_rangesn3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n3 = max(max_length_n3, curr_length)
					# For 4
					if x==4:
						curr_length +=1
						if curr_length>max_length_p4:
							count_p4=1
							time_rangesp4.clear()
							time_rangesp4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p4:
							count_p4+=1
							time_rangesp4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p4 = max(max_length_p4, curr_length)
					# For -4
					if x==-4:
						curr_length +=1
						if curr_length>max_length_n4:
							count_n4=1
							time_rangesn4.clear()
							time_rangesn4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n4:
							count_n4+=1
							time_rangesn4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n4 = max(max_length_n4, curr_length)

					# If j becomes last index, exited the loop by break statement because its work is done
					if j== size_a -1:
						break

					# Running while loop till the current Octant value is equal to next Octant value
					try:
						while df.loc[j,"Octant_value"]==df.loc[j+1,"Octant_value"]:
							# For 1
							if x==1:
								# Increasing current length by 1
								curr_length +=1
								# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
								if curr_length>max_length_p1:
									count_p1=1
									# Clearing time range if new max length found and appending that index to index range
									time_rangesp1.clear()
									time_rangesp1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								# If current length is equal to the maximum length then increase the count of maximum subsequence length
								elif curr_length==max_length_p1:
									count_p1+=1
									# Appending the list for from time to time as list to time range for current octant
									time_rangesp1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								# storing maximum subsequence length by comparing it with current length
								max_length_p1 = max(max_length_p1, curr_length)
							# For -1
							if x==-1:
								curr_length +=1
								if curr_length>max_length_n1:
									count_n1=1
									time_rangesn1.clear()
									time_rangesn1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n1:
									count_n1+=1
									time_rangesn1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n1 = max(max_length_n1, curr_length)
							# For 2
							if x==2:
								curr_length +=1
								if curr_length>max_length_p2:
									count_p2=1
									time_rangesp2.clear()
									time_rangesp2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p2:
									count_p2+=1
									time_rangesp2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p2 = max(max_length_p2, curr_length)
							# For -2
							if x==-2:
								curr_length +=1
								if curr_length>max_length_n2:
									count_n2=1
									time_rangesn2.clear()
									time_rangesn2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n2:
									count_n2+=1
									time_rangesn2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n2 = max(max_length_n2, curr_length)
							# For 3
							if x==3:
								curr_length +=1
								if curr_length>max_length_p3:
									count_p3=1
									time_rangesp3.clear()
									time_rangesp3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p3:
									count_p3+=1
									time_rangesp3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p3 = max(max_length_p3, curr_length)
							# For -3
							if x==-3:
								curr_length +=1
								if curr_length>max_length_n3:
									count_n3=1
									time_rangesn3.clear()
									time_rangesn3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n3:
									count_n3+=1
									time_rangesn3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n3 = max(max_length_n3, curr_length)
							# For 4
							if x==4:
								curr_length +=1
								if curr_length>max_length_p4:
									count_p4=1
									time_rangesp4.clear()
									time_rangesp4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p4:
									count_p4+=1
									time_rangesp4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p4 = max(max_length_p4, curr_length)
							# For -4
							if x==-4:
								curr_length +=1
								if curr_length>max_length_n4:
									count_n4=1
									time_rangesn4.clear()
									time_rangesn4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n4:
									count_n4+=1
									time_rangesn4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n4 = max(max_length_n4, curr_length)
							
							# Incrementing j by 1
							j = j+1

							# If j equals to last index after incrementing after while loop then we should compare
							# it with current octant or else it will lead to wrong output
							if j == size_a-1:
								# For 1
								if x==1:
									# Increasing current length by 1
									curr_length +=1
									# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
									if curr_length>max_length_p1:
										count_p1=1
										# Clearing time range if new max length found and appending that index to index range
										time_rangesp1.clear()
										time_rangesp1.append([df.loc[idx, "T"], df.loc[j , "T"]])
									# If current length is equal to the maximum length then increase the count of maximum subsequence length
									elif curr_length==max_length_p1:
										count_p1+=1
										# Appending the list for from time to time as list to time range for current octant
										time_rangesp1.append([df.loc[idx, "T"], df.loc[j , "T"]])
									# storing maximum subsequence length by comparing it with current length
									max_length_p1 = max(max_length_p1, curr_length)
								# For -1
								if x==-1:
									curr_length +=1
									if curr_length>max_length_n1:
										count_n1=1
										time_rangesn1.clear()
										time_rangesn1.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n1:
										count_n1+=1
										time_rangesn1.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n1 = max(max_length_n1, curr_length)
								# For 2
								if x==2:
									curr_length +=1
									if curr_length>max_length_p2:
										count_p2=1
										time_rangesp2.clear()
										time_rangesp2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p2:
										count_p2+=1
										time_rangesp2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p2 = max(max_length_p2, curr_length)
								# For -2
								if x==-2:
									curr_length +=1
									if curr_length>max_length_n2:
										count_n2=1
										time_rangesn2.clear()
										time_rangesn2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n2:
										count_n2+=1
										time_rangesn2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n2 = max(max_length_n2, curr_length)
								# For 3
								if x==3:
									curr_length +=1
									if curr_length>max_length_p3:
										count_p3=1
										time_rangesp3.clear()
										time_rangesp3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p3:
										count_p3+=1
										time_rangesp3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p3 = max(max_length_p3, curr_length)
								# For -3
								if x==-3:
									curr_length +=1
									if curr_length>max_length_n3:
										count_n3=1
										time_rangesn3.clear()
										time_rangesn3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n3:
										count_n3+=1
										time_rangesn3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n3 = max(max_length_n3, curr_length)
								# For 4
								if x==4:
									curr_length +=1
									if curr_length>max_length_p4:
										count_p4=1
										time_rangesp4.clear()
										time_rangesp4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p4:
										count_p4+=1
										time_rangesp4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p4 = max(max_length_p4, curr_length)
								# For -4
								if x==-4:
									curr_length +=1
									if curr_length>max_length_n4:
										count_n4=1
										time_rangesn4.clear()
										time_rangesn4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n4:
										count_n4+=1
										time_rangesn4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n4 = max(max_length_n4, curr_length)
								# Breaking the for loop for last index
								break
					
					except:
						print("Programme Failed :(")
						exit()
				
				# Finally storing max subsequence length of octant values at specified positions
				df.loc[0,"Longest Subsequence Length"]= max_length_p1
				df.loc[1,"Longest Subsequence Length"]= max_length_n1
				df.loc[2,"Longest Subsequence Length"]= max_length_p2
				df.loc[3,"Longest Subsequence Length"]= max_length_n2
				df.loc[4,"Longest Subsequence Length"]= max_length_p3
				df.loc[5,"Longest Subsequence Length"]= max_length_n3
				df.loc[6,"Longest Subsequence Length"]= max_length_p4
				df.loc[7,"Longest Subsequence Length"]= max_length_n4

				# Storing Count of each maximum length of octant values at specified locations in Count column
				df.loc[0,"Count"]= count_p1
				df.loc[1,"Count"]= count_n1
				df.loc[2,"Count"]= count_p2
				df.loc[3,"Count"]= count_n2
				df.loc[4,"Count"]= count_p3
				df.loc[5,"Count"]= count_n3
				df.loc[6,"Count"]= count_p4
				df.loc[7,"Count"]= count_n4

				# Creating columns for Storing time ranges
				
				df["_Octant_"]=np.nan
				df["Longest_Subsequence_Length"]=np.nan
				df["Count_"]=np.nan
				
				# Creating dictionary for Octant value
				dict_octant ={
					0 : "+1",
					1 : "-1",
					2 : "+2",
					3 : "-2",
					4 : "+3",
					5 : "-3",
					6 : "+4",
					7 : "-4",
				}
				# Creating dictionary for counts of longest subsequence length of octant values
				dict_count = {
					0: count_p1,
					1: count_n1,
					2: count_p2,
					3: count_n2,
					4: count_p3,
					5: count_n3,
					6: count_p4,
					7: count_n4,
				}
				# Creating dictionary for maximum length of subsequence of octant values
				dict_length = {
					0: max_length_p1,
					1: max_length_n1,
					2: max_length_p2,
					3: max_length_n2,
					4: max_length_p3,
					5: max_length_n3,
					6: max_length_p4,
					7: max_length_n4
				}
				# Creating dictionary for time ranges of octant values
				dict_time={
					0: time_rangesp1,
					1: time_rangesn1,
					2: time_rangesp2,
					3: time_rangesn2,
					4: time_rangesp3,
					5: time_rangesn3,
					6: time_rangesp4,
					7: time_rangesn4,

				}
				# Variable
				k =0

				# For loop for writing diffrent values of time ranges in dataframe
				for i in range(8):
					# Structure of output
					df.loc[k,"_Octant_"]= dict_octant[i]
					df.loc[k,"Longest_Subsequence_Length"]= dict_length[i]
					df.loc[k,"Count_"]= dict_count[i]
					df.loc[k+1,"_Octant_"]= "Time"
					df.loc[k+1,"Longest_Subsequence_Length"]= "From"
					df.loc[k+1,"Count_"]= "To"
					#Variable
					l=0
					# For loop for writing time ranges at specific position
					for item in dict_time[i]:
						df.loc[k+2+l,"Longest_Subsequence_Length"]=item[0]
						df.loc[k+2+l,"Count_"]=item[1]
						l=l+1
					# Incrementing k+2 by current count of current octant
					k= k+2 + dict_count[i]
				total_count =0
				for i in range(8):
					total_count +=dict_count[i]
				# pandas.io.formats.excel.ExcelFormatter.header_style = None
				# Saving dataframe to excel file
				# df.to_excel(path, index=False)
				# return total_count
				# Inserting empty columns
				df.insert(11,"",np.nan,True)
				df.insert(32,"",np.nan,True)
				df.insert(43,"",np.nan,True)
				df.insert(47,"",np.nan,True)
				# Renaming for avoiding confusion
				df.rename(columns = {'1_':''}, inplace = True)
				df.rename(columns = {'-1_':''}, inplace = True)
				df.rename(columns = {'2_':''}, inplace = True)
				df.rename(columns = {'-2_':''}, inplace = True)
				df.rename(columns = {'3_':''}, inplace = True)
				df.rename(columns = {'-3_':''}, inplace = True)
				df.rename(columns = {'4_':''}, inplace = True)
				df.rename(columns = {'-4_':''}, inplace = True)
				df.rename(columns = {'_':''}, inplace = True)
				
				pandas.io.formats.excel.ExcelFormatter.header_style = None
				# Saving dataframe to excel file
				cwd  = os.getcwd()
				directory = 'output'
				path  = os.path.join(cwd,directory)
				if not(os.path.exists(path)):
					os.mkdir(path)
				
				file_name= os.path.splitext(uploaded_file.name)[0]
				from datetime import datetime

				now = datetime.now()

				current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
				
				file_path = f'output\\{file_name}_'+f'{mod}_'+f'{current_time}.xlsx'
				file_namef = f'{file_name}_'+f'{mod}_'+f'{current_time}.xlsx'
				df.to_excel(file_path, index=False)
				# st.write(type(excel))
				# Opening workbook for manupulations
				wb= openpyxl.load_workbook(file_path)
				ws = wb['Sheet1']
				# Creating Border sides with black colour
				top = Side(border_style='thin', color="000000")
				bottom = Side(border_style='thin', color="000000")
				right = Side(border_style='thin', color="000000")
				left = Side(border_style='thin', color="000000")
				# Creating Border style
				border = Border(top=top, bottom = bottom, right=right, left=left)
				# Range for which border should be given
				range1=ws['N1':f'AF{iter+2}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				# Range for which border should be given
				range1 = ws[f'AC{iter+4}':f'AE{iter+12}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border

				# Range for which border should be given
				range1 = ws[f'AI3':f'AQ11']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				n = 0
				# While loop for creating border table
				while n<iter:
					range1 = ws[f'AI{17+13*n}':f'AQ{25+13*n}']
					for cell in range1:
						for x in cell:
							x.border=border
					n =n+1
				# Range for which border should be given
				range1 = ws[f'AS1':f'AU9']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				# Range for which border should be given
				range1 = ws[f'AW1':f'AY{total_count+17}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				
				# Creating style for highlighting cells and fonts
				fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
				font = Font(bold=False, color='000000')
				dxf=DifferentialStyle(font=font ,fill=fill_yellow)
				# Creating rule for cells
				rule = Rule(type='cellIs', operator='equal',formula=[1,1], dxf=dxf)
				# For loop for highlighting maximum transition count in row
				for i in range(8):
					range_cell = f'$AJ${4+i}:$AQ${4+i}'
					rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+range_cell+")"])
					ws.conditional_formatting.add(range_cell, rule1)

				# For loop for highlighting maximum transition count in row
				for m in range(iter):
					for i in range(8):
						range_cell = f'$AJ${18+13*m+i}:$AQ${18+13*m+i}'
						rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+range_cell+")"])
						ws.conditional_formatting.add(range_cell, rule1)
				
				# Creating stucture 
				ws.move_range(f"M1:AF{iter+12}", rows=1)
				ws.move_range(f"AS1:AU9", rows=2)
				ws.move_range(f"AW1:AY{17+total_count}", rows=2)
				ws['N1']="Overall Octant Count"
				ws['AS1']="Longest Subsequence Length"
				ws['AW1']="Longest Subsequence Length with Range"
				ws['M2']=''
				# For loop for highlighting Rank 1 cell
				for i in range(iter+1):
					ws.conditional_formatting.add(f'W{3+i}:AD{3+i}', rule)
				# Saving file to location
				wb.save(file_path)
				# Creating Downloadable file Button
				with open(file_path, 'rb') as my_file:

					st.write(f"Computed {uploaded_file.name}. Please download file from below button.")
					st.download_button(label = f'📥 Download {file_namef}', data = my_file, file_name = f'{file_namef}', mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 
				# st.download_button(label='📥 Download Current Result',data=excel ,file_name= 'df_test.xlsx')

	# If sidebar choosing Bulk file upload
	elif choice == "Bulk File Upload":
		# Subheader
		st.subheader("Bulk File Upload")
		# User Interface
		path1 = st.text_input('Enter the Folder Path for Bulk Conversion')
		mod = st.number_input('Please Enter a Mod input', value = 5000)
		# Compute Checkbox
		if st.checkbox("Compute"):
			os.chdir(path1)
			# Import required library for reading only excel files
			import ntpath
			ntpath.basename("a/b/c")
			# Function defination
			def path_leaf(path):
				head, tail = ntpath.split(path)
				return tail or ntpath.basename(head)
			import glob
			# Getting current directory
			# path1 = os.getcwd()
			# Creating lidt of excel files name
			file_list = glob.glob(path1+'\*.xlsx')
			# Creating list of files in directory
			dir_list = [path_leaf(path) for path in file_list]
			for uploaded_file in dir_list:
		
				# Creating variable for storing file name without extension
				file_name= os.path.splitext(uploaded_file)[0]
				
				try:
					# Copied content of input file into output file using shutil library and copyfile function
					# File is copied in target file, it is created if does not exist and overwrites if exists
					from datetime import datetime

					now = datetime.now()

					current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
					original = f'{path1}\{uploaded_file}'
					target = f'{path1}\{file_name}_{mod}_{current_time}.xlsx'
					shutil.copyfile(original, target)

					os.remove(f'{path1}\{uploaded_file}')
				# Except statement if error occured in try block like input file does not exist or Correct Directory is not open
				except EnvironmentError:
					print("Oops! Error Occured. Please Check Input file present in Directory and Correct name given to it.")
					print("Or Check if any file is open anywhere, close it and retry")
					print(f"Correct input file name: {uploaded_file}")
					exit()

				# Try block if output file do not created or exist
				try:
					# Created dataframe of output file using pandas library and read_excel() function 
					uploaded_file_form_name = f'{file_name}_{mod}_{current_time}.xlsx'
					df= pd.read_excel(uploaded_file_form_name)
					# size_a=df['U'].size
					size_a = 200
					
				# Except block if Output file is not found
				except:
					print("Output file not found. Error! Stoping the programme")
					exit()
				# st.write(uploaded_file)
				# Creating dictionary for octant name id mapping
				octant_name_id_mapping = {"1":"Internal outward interaction",
				"-1":"External outward interaction",
				"2":"External Ejection",
				"-2":"Internal Ejection",
				"3":"External inward interaction",
				"-3":"Internal inward interaction",
				"4":"Internal sweep",
				"-4":"External sweep"
				}
				# Creating dataframe from excel
				df = pd.read_excel(f'{uploaded_file_form_name}')
				
				# Calculating no. of iterations for mod ranges using ceil function of Math library
				try:
					iter = math.ceil(size_a/ mod)
					total_count=0
				# ZeroDivisionError for division by zero
				except ZeroDivisionError:
					print("Mod value cannot be zero! Please Change Mod value!!!")
					exit()
				# Created size variable and u_mean, v_mean, w_mean variables and stored means of individual columns in that
				
				u_mean=(df['U']).mean()
				v_mean=(df['V']).mean()
				w_mean=(df['W']).mean()

				# Try Block if Output file not found then to except
				try:
					#Created new blank columns for average velocities
					df["U avg"]= np.nan
					df["V avg"]= np.nan
					df["W avg"]= np.nan

					# Stored respective means at their position and column
					# .loc[] is used to access a group of rows and columns by labels
					df.loc[0,'U avg']=round(u_mean,3)
					df.loc[0,'V avg']=round(v_mean,3)
					df.loc[0,'W avg']=round(w_mean,3)

					# Created empty columns for U' = U- U avg
					# np.nan fills the column with NaN values
					df["U' = U - U avg"]= np.nan
					# Filled empty columns with required values of difference between U and U avg
					for i in range(0, size_a):
						df.loc[i,"U' = U - U avg"]= round(df.loc[i,'U']-u_mean,3)
					
					# Created empty columns for V' = V- V avg
					df["V' = V - V avg"]= np.nan
					# Filled empty columns with required values of difference between V and V avg
					for i in range(0, size_a):
						df.loc[i,"V' = V - V avg"]= round(df.loc[i,'V']-v_mean,3)

					# Created empty columns for W' = W- W avg
					df["W' = W - W avg"]= np.nan
					# Filled empty columns with required values of difference between W and W avg
					for i in range(0, size_a):
						df.loc[i,"W' = W - W avg"]= round(df.loc[i,'W']-w_mean,3)

					# Created an empty column for Octant Value storage
					df["Octant_value"]=np.nan
					
					# Logic for storing Octant Values
					# '''Iterated through all rows and compared different permutations and assigned Octant Value'''
					# Running Loop for every row
					for i in range(0, size_a):
						# Comparison for selecting octant value
						if df.loc[i,"U' = U - U avg"]>=0 and df.loc[i,"V' = V - V avg"]>=0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=1
							
							else:
								df.loc[i,"Octant_value"]=-1
						
						elif df.loc[i,"U' = U - U avg"]<0 and df.loc[i,"V' = V - V avg"]>=0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=2
							
							else:
								df.loc[i,"Octant_value"]=-2
						
						elif df.loc[i,"U' = U - U avg"]<0 and df.loc[i,"V' = V - V avg"]<0:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=3
							
							else:
								df.loc[i,"Octant_value"]=-3
						
						else:
							if df.loc[i,"W' = W - W avg"]>=0:
								df.loc[i,"Octant_value"]=4
							
							else:
								df.loc[i,"Octant_value"]=-4
					# st.write(df)
					# Typecasting Octant values from float to int
					df["Octant_value"]=df["Octant_value"].astype('Int64')
					
					# Created column
					df[""]=np.nan 
					
					# Just Created Good interphase for Excel file
					df.loc[0,"Octant ID"]="Overall Octant Count"

					# Initialized count variables with zero   
					count_pos1=0
					count_neg1=0
					count_pos2=0
					count_neg2=0
					count_pos3=0
					count_neg3=0
					count_pos4=0
					count_neg4=0
						
					# Logic for counting octant values
					# Iterated through Octant Value column and compared with octant value and increased count by one        
					
					for i in range(0,size_a):
						# print(type(df.loc[i,"Octant_value"]))
						if df.loc[i,"Octant_value"]==1:
							count_pos1 +=1
						elif df.loc[i,"Octant_value"]==-1:
							count_neg1 +=1
						elif df.loc[i,"Octant_value"]==2:
							count_pos2 +=1
						elif df.loc[i,"Octant_value"]==-2:
							count_neg2 +=1
						elif df.loc[i,"Octant_value"]==3:
							count_pos3 +=1
						elif df.loc[i,"Octant_value"]==-3:
							count_neg3 +=1
						elif df.loc[i,"Octant_value"]==4:
							count_pos4 +=1
						elif df.loc[i, "Octant_value"]==-4:
							count_neg4 +=1
				
					# Written values of counts at specified positions
					df.loc[0,"1"]=count_pos1
					df.loc[0,"-1"]=count_neg1
					df.loc[0,"2"]=count_pos2
					df.loc[0,"-2"]=count_neg2
					df.loc[0,"3"]=count_pos3
					df.loc[0,"-3"]=count_neg3        
					df.loc[0,"4"]=count_pos4   
					df.loc[0,"-4"]=count_neg4
					
					# Dictionary for mapping index to octant values
					util={
						1:1,
						2:-1,
						3:2,
						4:-2,
						5:3,
						6:-3,
						7:4,
						8:-4
					}
					# Logic: Creating a list for count values and creaating a copy of it. Sorting the copy of list.
					# Creating Dictionary of copy list and storing values as keys, indices as values
					ls = [count_pos1, count_neg1, count_pos2, count_neg2, count_pos3, count_neg3, count_pos4, count_neg4]
					ls_ = ls.copy()
					ls_.sort(reverse=True)
					ls_dict = {k: v for v, k in enumerate(ls_)}
					
					# Then Storing Ranks at specified positions
					for m in range(8):
						df.loc[0,f'Rank Octant {util[m+1]}']=ls_dict[ls[m]]+1
						if ls_dict[ls[m]]==0:
							store=m
					# Creating Structure and storing Rank 1's Octant ID and its name
					df.loc[0,"Rank 1 Octant ID"] = util[store+1]
					df.loc[0, "Rank 1 Octant Name"] = octant_name_id_mapping[str(util[store+1])]
					
					s = "Mod "+ str(mod)
					df.loc[0,""]= s

					
					# Creating a list for storing mod Ranks
					ranklist=[]
					# Iterating through each range and count the octant value of each ID
					for i in range(2, iter+2):
						# Writing range in file
						stg=""
						if i==2:
							x = min(mod*(i-1)-1,size_a)
							stg = "0000"+"-"+str(x)
						elif i== iter+1:
							stg= str(mod*(i-2))+"-"+str(size_a-1)
						else:
							stg = str(mod *(i-2))+"-"+str(mod*(i-1)-1)
						df.loc[i-1,"Octant ID"]= stg

						# Initializing mod counts as zeroes
						mod_cnt_pos_one=0
						mod_cnt_neg_one=0
						mod_cnt_pos_two=0
						mod_cnt_neg_two=0
						mod_cnt_pos_three=0
						mod_cnt_neg_three=0
						mod_cnt_pos_four=0
						mod_cnt_neg_four=0

						# Handling current range for last call
						current_range= min(size_a, (i-1)*mod)

						# Iterating over range and incrementing count for respective mod count
						for j in range((i-2)*mod, current_range):
							if df.loc[j, "Octant_value"]==1:
								mod_cnt_pos_one+=1
							elif df.loc[j, "Octant_value"]==-1:
								mod_cnt_neg_one+=1
							elif df.loc[j, "Octant_value"]==2:
								mod_cnt_pos_two+=1
							elif df.loc[j, "Octant_value"]==-2:
								mod_cnt_neg_two+=1
							elif df.loc[j, "Octant_value"]==3:
								mod_cnt_pos_three+=1
							elif df.loc[j, "Octant_value"]==-3:
								mod_cnt_neg_three+=1
							elif df.loc[j, "Octant_value"]==4:
								mod_cnt_pos_four+=1
							elif df.loc[j, "Octant_value"]==-4:
								mod_cnt_neg_four+=1
						
						# writing mod counts at specified locations
						df.loc[i-1, "1"]=mod_cnt_pos_one
						df.loc[i-1, "-1"]=mod_cnt_neg_one
						df.loc[i-1, "2"]=mod_cnt_pos_two
						df.loc[i-1, "-2"]=mod_cnt_neg_two
						df.loc[i-1, "3"]=mod_cnt_pos_three
						df.loc[i-1, "-3"]=mod_cnt_neg_three
						df.loc[i-1, "4"]=mod_cnt_pos_four
						df.loc[i-1, "-4"]=mod_cnt_neg_four
						
						# Logic: Creating a list for count values and creaating a copy of it. Sorting the copy of list.
						# Creating Dictionary of copy list and storing values as keys, indices as values
						ls = [mod_cnt_pos_one, mod_cnt_neg_one, mod_cnt_pos_two, mod_cnt_neg_two, mod_cnt_pos_three, mod_cnt_neg_three, mod_cnt_pos_four, mod_cnt_neg_four]
						ls_ = ls.copy()
						ls_.sort(reverse=True)
						ls_dict = {k: v for v, k in enumerate(ls_)}
						# Then Storing Ranks at specified positions
						for m in range(8):
							df.loc[i-1,f'Rank Octant {util[m+1]}']=ls_dict[ls[m]]+1
							if ls_dict[ls[m]]==0:
								store=m
						# Creating structure and storing Rank 1's Octant ID
						df.loc[i-1,"Rank 1 Octant ID"] = util[store+1]
						# Appending current Rank's octant to ranklist
						ranklist.append(util[store+1])
						# Storing current Octant's name
						df.loc[i-1, "Rank 1 Octant Name"] = octant_name_id_mapping[str(util[store+1])]
					
					# Creating structure 
					df.loc[iter+2,"Rank Octant 4"]="Octant ID"
					df.loc[iter+2,"Rank Octant -4"]="Octant Name"
					df.loc[iter+2,"Rank 1 Octant ID"]="Count of Rank 1 Mod Values"
					# Iterating over loop and storing mod counts of each octant
					for n in range(iter+3, iter+11):
						df.loc[n,"Rank Octant 4"]= (util[n-iter-2])
						df.loc[n,"Rank Octant -4"]=octant_name_id_mapping[str(util[n-iter-2])]
						df.loc[n,"Rank 1 Octant ID"]= ranklist.count(util[n-iter-2])
					
					

				# Except block if File not found or incorrect directory is opened
				except FileNotFoundError:
					st.write("Oops! Error Occured! File Not Found!")
					st.write("Please Check File present in correct directory with correct name")
					exit()

				#Creating Structure for Overall Transition count Table
				df["_"]=np.nan
				df["Overall Transition Count"]=np.nan
				df.loc[1,"Overall Transition Count"]="Octant"
				df.loc[0,"1_"]="To"
				df.loc[1,"_"]="From"
				df.loc[2,"Overall Transition Count"]="+1"
				df.loc[3,"Overall Transition Count"]="-1"
				df.loc[4,"Overall Transition Count"]="+2"
				df.loc[5,"Overall Transition Count"]="-2"
				df.loc[6,"Overall Transition Count"]="+3"
				df.loc[7,"Overall Transition Count"]="-3"
				df.loc[8,"Overall Transition Count"]="+4"
				df.loc[9,"Overall Transition Count"]="-4"

				df.loc[1,"1_"]="+1"
				df.loc[1,"-1_"]="-1"
				df.loc[1,"2_"]="+2"
				df.loc[1,"-2_"]="-2"
				df.loc[1,"3_"]="+3"
				df.loc[1,"-3_"]="-3"
				df.loc[1,"4_"]="+4"
				df.loc[1,"-4_"]="-4"

				# Initializing every cell in table as 0
				for i in range(8):
					df.loc[2+i,"1_"]=0
					df.loc[2+i,"-1_"]=0
					df.loc[2+i,"2_"]=0
					df.loc[2+i,"-2_"]=0
					df.loc[2+i,"3_"]=0
					df.loc[2+i,"-3_"]=0
					df.loc[2+i,"4_"]=0
					df.loc[2+i,"-4_"]=0


				# Counting and filling the table of overall transition count
				for i in range(0,size_a-1):
					x= df.loc[i,"Octant_value"]
					y=df.loc[i+1,"Octant_value"]
					z=0
					if x>0:
						z= x*2 -1
					elif x<0:
						z= -2 * x
					
					df.loc[1+z,f'{y}_']=df.loc[1+z, f'{y}_']+1

				# Counting the mod ranges using this for loop
				for i in range(iter):
					# Creating structures for mod ranges count
					df.loc[iter+19+13*i-iter-6, "Overall Transition Count"]="Mod Transition Count"
					df.loc[iter+21+13*i-iter-6,"Overall Transition Count"]="Octant"
					df.loc[iter+22+13*i-iter-6,"_"]="From"
					df.loc[iter+20+13*i-iter-6,"1_"]="To"
					df.loc[iter+22+13*i-iter-6,"Overall Transition Count"]="+1"
					df.loc[iter+23+13*i-iter-6,"Overall Transition Count"]="-1"
					df.loc[iter+24+13*i-iter-6,"Overall Transition Count"]="+2"
					df.loc[iter+25+13*i-iter-6,"Overall Transition Count"]="-2"
					df.loc[iter+26+13*i-iter-6,"Overall Transition Count"]="+3"
					df.loc[iter+27+13*i-iter-6,"Overall Transition Count"]="-3"
					df.loc[iter+28+13*i-iter-6,"Overall Transition Count"]="+4"
					df.loc[iter+29+13*i-iter-6,"Overall Transition Count"]="-4"

					df.loc[iter+21+13*i-iter-6,"1_"]="+1"
					df.loc[iter+21+13*i-iter-6,"-1_"]="-1"
					df.loc[iter+21+13*i-iter-6,"2_"]="+2"
					df.loc[iter+21+13*i-iter-6,"-2_"]="-2"
					df.loc[iter+21+13*i-iter-6,"3_"]="+3"
					df.loc[iter+21+13*i-iter-6,"-3_"]="-3"
					df.loc[iter+21+13*i-iter-6,"4_"]="+4"
					df.loc[iter+21+13*i-iter-6,"-4_"]="-4"

					# Initializing every cell of table to zero
					for j in range(8):
						df.loc[iter+22+j+13*i-iter-6,"1_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-1_"]=0
						df.loc[iter+22+j+13*i-iter-6,"2_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-2_"]=0
						df.loc[iter+22+j+13*i-iter-6,"3_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-3_"]=0
						df.loc[iter+22+j+13*i-iter-6,"4_"]=0
						df.loc[iter+22+j+13*i-iter-6,"-4_"]=0

					# Initializing start and end variables
					start = i * mod
					end = ((i+1) * mod) if ((i+1) * mod) < size_a else size_a-1
					# Creating string for mod ranges and storing at correct location
					if i==0:
						x = min(mod*(i+1),size_a)
						s = "0000"+"-"+str(x-1)
					elif i== iter-1:
						s= str(start)+"-"+str(size_a-1)
					else:
						s = str(start)+"-"+str(end-1)
					df.loc[iter+20+13*i-iter-6,"Overall Transition Count"]= s
					
					#Storing count of transition in cell of table after iterating through range 
					for k in range(start,end):
						x= df.loc[k,"Octant_value"]
						y= df.loc[k+1,"Octant_value"]
						z=0
						if x>0:
							z= x*2 -1
						elif x<0:
							z= -2 * x
						
						df.loc[iter+21+13*i+z-iter-6,f'{y}_']=df.loc[iter+21+z+13*i-iter-6, f'{y}_']+1
				
				df["Octant"]=np.nan
				df.loc[0, "Octant"]="+1"
				df.loc[1, "Octant"]="-1"
				df.loc[2, "Octant"]="+2"
				df.loc[3, "Octant"]="-2"
				df.loc[4, "Octant"]="+3"
				df.loc[5, "Octant"]="-3"
				df.loc[6, "Octant"]="+4"
				df.loc[7, "Octant"]="-4"

				# Created columns Longest subsequence length and count
				df["Longest Subsequence Length"]=np.nan
				df["Count"]=np.nan

				# Initialized variables for max length of subsequence of each octant
				max_length_p1=0
				max_length_n1=0
				max_length_p2=0
				max_length_n2=0
				max_length_p3=0
				max_length_n3=0
				max_length_p4=0
				max_length_n4=0

				# Initialized Count variables for each octant value
				count_p1 = 0
				count_n1 = 0
				count_p2 = 0
				count_n2 = 0
				count_p3 = 0
				count_n3 = 0
				count_p4 = 0
				count_n4 = 0
				# Initialized time range variables of octant as list to store j
				time_rangesp1 = []
				time_rangesn1 = []
				time_rangesp2 = []
				time_rangesn2 = []
				time_rangesp3 = []
				time_rangesn3 = []
				time_rangesp4 = []
				time_rangesn4 = []


				# Iterated through overall range of octant columns
				for j in range(size_a):
					# Initialized curr_length variable as 0 for keep track of length of current octant value subsequence
					curr_length = 0
					# Variable x for storing Octant Value
					x = df.loc[j,"Octant_value"]
					# Current index as idx
					idx = j

					# Comparing current octant value with each 8 octant values and incrementing corresponding 
					# max subsequence length variable and count of that max subsequence
					# For 1
					if x==1:
						# Increasing current length by 1
						curr_length +=1
						# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
						if curr_length>max_length_p1:
							count_p1=1
							# Clearing time range if new max length found and appending that index to index range
							time_rangesp1.clear()
							time_rangesp1.append([df.loc[idx, "T"], df.loc[idx , "T"]])
						# If current length is equal to the maximum length then increase the count of maximum subsequence length
						elif curr_length==max_length_p1:
							count_p1+=1
							# Appending the list for from time to time as list to time range for current octant
							time_rangesp1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						# storing maximum subsequence length by comparing it with current length
						max_length_p1 = max(max_length_p1, curr_length)
					
					# For -1
					if x==-1:
						curr_length +=1
						if curr_length>max_length_n1:
							count_n1=1
							time_rangesn1.clear()
							time_rangesn1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n1:
							count_n1+=1
							time_rangesn1.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n1 = max(max_length_n1, curr_length)

					# For 2
					if x==2:
						curr_length +=1
						if curr_length>max_length_p2:
							count_p2=1
							time_rangesp2.clear()
							time_rangesp2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p2:
							count_p2+=1
							time_rangesp2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p2 = max(max_length_p2, curr_length)
					# For -2
					if x==-2:
						curr_length +=1
						if curr_length>max_length_n2:
							count_n2=1
							time_rangesn2.clear()
							time_rangesn2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n2:
							count_n2+=1
							time_rangesn2.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n2 = max(max_length_n2, curr_length)
					# For 3
					if x==3:
						curr_length +=1
						if curr_length>max_length_p3:
							count_p3=1
							time_rangesp3.clear()
							time_rangesp3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p3:
							count_p3+=1
							time_rangesp3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p3 = max(max_length_p3, curr_length)
					# For -3
					if x==-3:
						curr_length +=1
						if curr_length>max_length_n3:
							count_n3=1
							time_rangesn3.clear()
							time_rangesn3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n3:
							count_n3+=1
							time_rangesn3.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n3 = max(max_length_n3, curr_length)
					# For 4
					if x==4:
						curr_length +=1
						if curr_length>max_length_p4:
							count_p4=1
							time_rangesp4.clear()
							time_rangesp4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_p4:
							count_p4+=1
							time_rangesp4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_p4 = max(max_length_p4, curr_length)
					# For -4
					if x==-4:
						curr_length +=1
						if curr_length>max_length_n4:
							count_n4=1
							time_rangesn4.clear()
							time_rangesn4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						elif curr_length==max_length_n4:
							count_n4+=1
							time_rangesn4.append([df.loc[idx, "T"], df.loc[idx, "T"]])
						max_length_n4 = max(max_length_n4, curr_length)

					# If j becomes last index, exited the loop by break statement because its work is done
					if j== size_a -1:
						break

					# Running while loop till the current Octant value is equal to next Octant value
					try:
						while df.loc[j,"Octant_value"]==df.loc[j+1,"Octant_value"]:
							# For 1
							if x==1:
								# Increasing current length by 1
								curr_length +=1
								# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
								if curr_length>max_length_p1:
									count_p1=1
									# Clearing time range if new max length found and appending that index to index range
									time_rangesp1.clear()
									time_rangesp1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								# If current length is equal to the maximum length then increase the count of maximum subsequence length
								elif curr_length==max_length_p1:
									count_p1+=1
									# Appending the list for from time to time as list to time range for current octant
									time_rangesp1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								# storing maximum subsequence length by comparing it with current length
								max_length_p1 = max(max_length_p1, curr_length)
							# For -1
							if x==-1:
								curr_length +=1
								if curr_length>max_length_n1:
									count_n1=1
									time_rangesn1.clear()
									time_rangesn1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n1:
									count_n1+=1
									time_rangesn1.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n1 = max(max_length_n1, curr_length)
							# For 2
							if x==2:
								curr_length +=1
								if curr_length>max_length_p2:
									count_p2=1
									time_rangesp2.clear()
									time_rangesp2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p2:
									count_p2+=1
									time_rangesp2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p2 = max(max_length_p2, curr_length)
							# For -2
							if x==-2:
								curr_length +=1
								if curr_length>max_length_n2:
									count_n2=1
									time_rangesn2.clear()
									time_rangesn2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n2:
									count_n2+=1
									time_rangesn2.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n2 = max(max_length_n2, curr_length)
							# For 3
							if x==3:
								curr_length +=1
								if curr_length>max_length_p3:
									count_p3=1
									time_rangesp3.clear()
									time_rangesp3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p3:
									count_p3+=1
									time_rangesp3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p3 = max(max_length_p3, curr_length)
							# For -3
							if x==-3:
								curr_length +=1
								if curr_length>max_length_n3:
									count_n3=1
									time_rangesn3.clear()
									time_rangesn3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n3:
									count_n3+=1
									time_rangesn3.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n3 = max(max_length_n3, curr_length)
							# For 4
							if x==4:
								curr_length +=1
								if curr_length>max_length_p4:
									count_p4=1
									time_rangesp4.clear()
									time_rangesp4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_p4:
									count_p4+=1
									time_rangesp4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_p4 = max(max_length_p4, curr_length)
							# For -4
							if x==-4:
								curr_length +=1
								if curr_length>max_length_n4:
									count_n4=1
									time_rangesn4.clear()
									time_rangesn4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								elif curr_length==max_length_n4:
									count_n4+=1
									time_rangesn4.append([df.loc[idx, "T"], df.loc[j+1, "T"]])
								max_length_n4 = max(max_length_n4, curr_length)
							
							# Incrementing j by 1
							j = j+1

							# If j equals to last index after incrementing after while loop then we should compare
							# it with current octant or else it will lead to wrong output
							if j == size_a-1:
								# For 1
								if x==1:
									# Increasing current length by 1
									curr_length +=1
									# Checking if current length is greater than max subsequence length till now then count of that octant will become 1
									if curr_length>max_length_p1:
										count_p1=1
										# Clearing time range if new max length found and appending that index to index range
										time_rangesp1.clear()
										time_rangesp1.append([df.loc[idx, "T"], df.loc[j , "T"]])
									# If current length is equal to the maximum length then increase the count of maximum subsequence length
									elif curr_length==max_length_p1:
										count_p1+=1
										# Appending the list for from time to time as list to time range for current octant
										time_rangesp1.append([df.loc[idx, "T"], df.loc[j , "T"]])
									# storing maximum subsequence length by comparing it with current length
									max_length_p1 = max(max_length_p1, curr_length)
								# For -1
								if x==-1:
									curr_length +=1
									if curr_length>max_length_n1:
										count_n1=1
										time_rangesn1.clear()
										time_rangesn1.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n1:
										count_n1+=1
										time_rangesn1.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n1 = max(max_length_n1, curr_length)
								# For 2
								if x==2:
									curr_length +=1
									if curr_length>max_length_p2:
										count_p2=1
										time_rangesp2.clear()
										time_rangesp2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p2:
										count_p2+=1
										time_rangesp2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p2 = max(max_length_p2, curr_length)
								# For -2
								if x==-2:
									curr_length +=1
									if curr_length>max_length_n2:
										count_n2=1
										time_rangesn2.clear()
										time_rangesn2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n2:
										count_n2+=1
										time_rangesn2.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n2 = max(max_length_n2, curr_length)
								# For 3
								if x==3:
									curr_length +=1
									if curr_length>max_length_p3:
										count_p3=1
										time_rangesp3.clear()
										time_rangesp3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p3:
										count_p3+=1
										time_rangesp3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p3 = max(max_length_p3, curr_length)
								# For -3
								if x==-3:
									curr_length +=1
									if curr_length>max_length_n3:
										count_n3=1
										time_rangesn3.clear()
										time_rangesn3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n3:
										count_n3+=1
										time_rangesn3.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n3 = max(max_length_n3, curr_length)
								# For 4
								if x==4:
									curr_length +=1
									if curr_length>max_length_p4:
										count_p4=1
										time_rangesp4.clear()
										time_rangesp4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_p4:
										count_p4+=1
										time_rangesp4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_p4 = max(max_length_p4, curr_length)
								# For -4
								if x==-4:
									curr_length +=1
									if curr_length>max_length_n4:
										count_n4=1
										time_rangesn4.clear()
										time_rangesn4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									elif curr_length==max_length_n4:
										count_n4+=1
										time_rangesn4.append([df.loc[idx, "T"], df.loc[j, "T"]])
									max_length_n4 = max(max_length_n4, curr_length)
								# Breaking the for loop for last index
								break
					
					except:
						print("Programme Failed :(")
						exit()
				
				# Finally storing max subsequence length of octant values at specified positions
				df.loc[0,"Longest Subsequence Length"]= max_length_p1
				df.loc[1,"Longest Subsequence Length"]= max_length_n1
				df.loc[2,"Longest Subsequence Length"]= max_length_p2
				df.loc[3,"Longest Subsequence Length"]= max_length_n2
				df.loc[4,"Longest Subsequence Length"]= max_length_p3
				df.loc[5,"Longest Subsequence Length"]= max_length_n3
				df.loc[6,"Longest Subsequence Length"]= max_length_p4
				df.loc[7,"Longest Subsequence Length"]= max_length_n4

				# Storing Count of each maximum length of octant values at specified locations in Count column
				df.loc[0,"Count"]= count_p1
				df.loc[1,"Count"]= count_n1
				df.loc[2,"Count"]= count_p2
				df.loc[3,"Count"]= count_n2
				df.loc[4,"Count"]= count_p3
				df.loc[5,"Count"]= count_n3
				df.loc[6,"Count"]= count_p4
				df.loc[7,"Count"]= count_n4

				# Creating columns for Storing time ranges
				
				df["_Octant_"]=np.nan
				df["Longest_Subsequence_Length"]=np.nan
				df["Count_"]=np.nan
				
				# Creating dictionary for Octant value
				dict_octant ={
					0 : "+1",
					1 : "-1",
					2 : "+2",
					3 : "-2",
					4 : "+3",
					5 : "-3",
					6 : "+4",
					7 : "-4",
				}
				# Creating dictionary for counts of longest subsequence length of octant values
				dict_count = {
					0: count_p1,
					1: count_n1,
					2: count_p2,
					3: count_n2,
					4: count_p3,
					5: count_n3,
					6: count_p4,
					7: count_n4,
				}
				# Creating dictionary for maximum length of subsequence of octant values
				dict_length = {
					0: max_length_p1,
					1: max_length_n1,
					2: max_length_p2,
					3: max_length_n2,
					4: max_length_p3,
					5: max_length_n3,
					6: max_length_p4,
					7: max_length_n4
				}
				# Creating dictionary for time ranges of octant values
				dict_time={
					0: time_rangesp1,
					1: time_rangesn1,
					2: time_rangesp2,
					3: time_rangesn2,
					4: time_rangesp3,
					5: time_rangesn3,
					6: time_rangesp4,
					7: time_rangesn4,

				}
				# Variable
				k =0

				# For loop for writing diffrent values of time ranges in dataframe
				for i in range(8):
					# Structure of output
					df.loc[k,"_Octant_"]= dict_octant[i]
					df.loc[k,"Longest_Subsequence_Length"]= dict_length[i]
					df.loc[k,"Count_"]= dict_count[i]
					df.loc[k+1,"_Octant_"]= "Time"
					df.loc[k+1,"Longest_Subsequence_Length"]= "From"
					df.loc[k+1,"Count_"]= "To"
					#Variable
					l=0
					# For loop for writing time ranges at specific position
					for item in dict_time[i]:
						df.loc[k+2+l,"Longest_Subsequence_Length"]=item[0]
						df.loc[k+2+l,"Count_"]=item[1]
						l=l+1
					# Incrementing k+2 by current count of current octant
					k= k+2 + dict_count[i]
				total_count =0
				for i in range(8):
					total_count +=dict_count[i]
				# pandas.io.formats.excel.ExcelFormatter.header_style = None
				# Saving dataframe to excel file
				# df.to_excel(path, index=False)
				# return total_count
				# Inserting empty columns
				df.insert(11,"",np.nan,True)
				df.insert(32,"",np.nan,True)
				df.insert(43,"",np.nan,True)
				df.insert(47,"",np.nan,True)
				# Renaming for avoiding confusion
				df.rename(columns = {'1_':''}, inplace = True)
				df.rename(columns = {'-1_':''}, inplace = True)
				df.rename(columns = {'2_':''}, inplace = True)
				df.rename(columns = {'-2_':''}, inplace = True)
				df.rename(columns = {'3_':''}, inplace = True)
				df.rename(columns = {'-3_':''}, inplace = True)
				df.rename(columns = {'4_':''}, inplace = True)
				df.rename(columns = {'-4_':''}, inplace = True)
				df.rename(columns = {'_':''}, inplace = True)
				
				pandas.io.formats.excel.ExcelFormatter.header_style = None
				# Saving dataframe to excel file
				# cwd  = os.getcwd()
				# directory = 'output'
				# path  = os.path.join(cwd,directory)
				# if not(os.path.exists(path)):
				# 	os.mkdir(path)
				
				# file_name= os.path.splitext(uploaded_file.name)[0]
				from datetime import datetime

				now = datetime.now()

				# current_time = now.strftime("%Y-%m-%d-%H-%M-%S")
				
				# file_path = f'{file_name}_'+f'{mod}_'+f'{current_time}.xlsx'
				# file_namef = f'{file_name}_'+f'{mod}_'+f'{current_time}.xlsx'
				df.to_excel(uploaded_file_form_name, index=False)
				# st.write(type(excel))
				# Opening workbook for manupulations
				wb= openpyxl.load_workbook(uploaded_file_form_name)
				ws = wb['Sheet1']
				# Creating Border sides with black colour
				top = Side(border_style='thin', color="000000")
				bottom = Side(border_style='thin', color="000000")
				right = Side(border_style='thin', color="000000")
				left = Side(border_style='thin', color="000000")
				# Creating Border style
				border = Border(top=top, bottom = bottom, right=right, left=left)
				# Range for which border should be given
				range1=ws['N1':f'AF{iter+2}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				# Range for which border should be given
				range1 = ws[f'AC{iter+4}':f'AE{iter+12}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border

				# Range for which border should be given
				range1 = ws[f'AI3':f'AQ11']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				n = 0
				# While loop for creating border table
				while n<iter:
					range1 = ws[f'AI{17+13*n}':f'AQ{25+13*n}']
					for cell in range1:
						for x in cell:
							x.border=border
					n =n+1
				# Range for which border should be given
				range1 = ws[f'AS1':f'AU9']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				# Range for which border should be given
				range1 = ws[f'AW1':f'AY{total_count+17}']
				# For loop for creating border table
				for cell in range1:
					for x in cell:
						x.border=border
				
				# Creating style for highlighting cells and fonts
				fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
				font = Font(bold=False, color='000000')
				dxf=DifferentialStyle(font=font ,fill=fill_yellow)
				# Creating rule for cells
				rule = Rule(type='cellIs', operator='equal',formula=[1,1], dxf=dxf)
				# For loop for highlighting maximum transition count in row
				for i in range(8):
					range_cell = f'$AJ${4+i}:$AQ${4+i}'
					rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+range_cell+")"])
					ws.conditional_formatting.add(range_cell, rule1)

				# For loop for highlighting maximum transition count in row
				for m in range(iter):
					for i in range(8):
						range_cell = f'$AJ${18+13*m+i}:$AQ${18+13*m+i}'
						rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+range_cell+")"])
						ws.conditional_formatting.add(range_cell, rule1)
				
				# Creating stucture 
				ws.move_range(f"M1:AF{iter+12}", rows=1)
				ws.move_range(f"AS1:AU9", rows=2)
				ws.move_range(f"AW1:AY{17+total_count}", rows=2)
				ws['N1']="Overall Octant Count"
				ws['AS1']="Longest Subsequence Length"
				ws['AW1']="Longest Subsequence Length with Range"
				ws['M2']=''
				# For loop for highlighting Rank 1 cell
				for i in range(iter+1):
					ws.conditional_formatting.add(f'W{3+i}:AD{3+i}', rule)
				# Saving file to location
				wb.save(uploaded_file_form_name)
				with open(uploaded_file_form_name, 'rb') as my_file:
					# Compute Button for Download
					st.write(f"Computed {uploaded_file_form_name}. Please download file from below button.")
					st.download_button(label = f'Download {uploaded_file_form_name}', data = my_file, file_name = f'{uploaded_file_form_name}', mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 
				
# Function Call
proj_octant_gui()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
