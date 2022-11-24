import os
os.system("cls")
import pandas as pd
import numpy as np
import openpyxl
import math
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border,Side

from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
import concurrent.futures
#importing libraries

from datetime import datetime
start_time = datetime.now()
os.chdir(r'C:\Users\pc\Documents\GitHub\2001CE45_2022\tut07')

#Help
def check_octant(x,y,z):
    #this function will check for the octant
        #first it will check for quadrant
        #then it will update it for the octant by looking the z values(w-w_avg for the main data)
    if x>=0 and y>=0 :
        o=1
    if x<0 and y>=0 :
        o=2
    if x<0 and y<0 :
        o=3
    if x>=0 and y<0 :
        o=4
        #till here quqdrant is checked
    if z>=0 :
        return o
    if z<0 :
        return o*(-1)
        #octant value returned as o
def octant_analysis(file, mod=5000):
	#readymadefunction having argument as mod value
    #complete code has been written in this(major portion)
    try:
        datain =  pd.read_excel("input\\"+file) #reading input csv file and storing in variable datain as data input
        total_size=datain['U'].size  #size stored in variable
        U_Avg=datain['U'].mean()
        V_Avg=datain['V'].mean()
        W_Avg=datain['W'].mean()
        #mean value for each velocity value stored in variable
        datain['U_Avg']=''
        datain['V_Avg']=np.nan
        datain['W_Avg']=None
        #blank column naming average has been created in the three different ways
        datain.loc[0,'U_Avg']=round(U_Avg,3)
        datain.loc[0,'V_Avg']=round(V_Avg,3)
        datain.loc[0,'W_Avg']=round(W_Avg,3)
        #storing the values of avg valocity in 1st row or 0 indexed row
        datain['U-U_Avg']=round(datain['U']-U_Avg, 3)
        datain['V-V_Avg']=round(datain['V']-V_Avg, 3)
        datain['W-W_Avg']=round(datain['W']-W_Avg, 3)
        #three more columns have been made and assigned value as V-Vavg.....corresponding name and corresponding columns
        datain['Octant']=np.nan #created a blank column named octant for which operation is to be done
        datain['']=np.nan
        i=0 #initiating for loop variable
        while i<total_size: #loop is run for whole column size of input column file
            datain.loc[i,'Octant']=check_octant(datain.loc[i,'U-U_Avg'],datain.loc[i,'V-V_Avg'],datain.loc[i,'W-W_Avg'])
            #octant column is assigned octant value after chehking value of octant by calling function
            i=i+1 #updating loop

        datain['remove_mod']=np.nan #new blank file and some processing according to the demand of output file has been done
        datain.loc[0,'remove_mod']="mod {}".format(mod) 
        #assigned a string at given index 1 acc. to output file
        datain['Octant Id']=np.nan 
        datain['1']=np.nan
        datain['-1']=np.nan
        datain['2']=np.nan
        datain['-2']=np.nan
        datain['3']=np.nan
        datain['-3']=np.nan
        datain['4']=np.nan
        datain['-4']=np.nan

        #new columns has been made for further proceedings
        datain.loc[0,'Octant Id']="Overall Count"
        #again filling the column but this time adding parameters in it on which the next processing depends
        noi = math.ceil(total_size/mod) #here ceil value provides the upper bound of the floating no...e.g..7.4 have ceil value 8
        #this will reflect no. of iterations to be performed or no. of sections in which our input file will be divided to check the count 
        i=0 #initiated for iteration
        op1=0
        on1=0
        op2=0
        on2=0
        op3=0
        on3=0
        op4=0
        on4=0
        #initiated to count overall count of each octant op1-overall positive 1,on1-overall negative 1
        while i<noi: #loop conditions
            if(i+1)*mod>total_size: 
                u=total_size-1 
            else:             #u is the condition or limit for upper boundary of the mod ranged values
                u=(i+1)*mod-1
            l=i*mod          #similarly l is the lower boundary
            datain.loc[i+1,'Octant Id']="{}-{}".format(l,u) #filling mod ranged section as its demand of sections should not count more than mod value count columns
            j=0 #initiated for loop
            p1=0
            n1=0
            p2=0
            n2=0
            p3=0
            n3=0
            p4=0
            n4=0
            #these values initiated for filling frequency of occurence of octant in each mod range 
            #p1--positive 1.......n1--negative 1

            #this is loop is being run to count the each octant values in each column range max of mod and and calculate total occurence of each occurence in the column range

            while j<mod: #loop condition
                
                pos =(i*mod)+j #variable to get the column no. where exactly we are counting the octant value
                if pos>=total_size:
                    break #conditions fulfilling criteria of loop
                else:
                    if datain.loc[pos,'Octant']==1:
                        p1=p1+1
                    elif datain.loc[pos,'Octant']==-1:
                        n1=n1+1
                    elif datain.loc[pos,'Octant']==2:
                        p2=p2+1
                    elif datain.loc[pos,'Octant']==-2:
                        n2=n2+1
                    elif datain.loc[pos,'Octant']==3:
                        p3=p3+1
                    elif datain.loc[pos,'Octant']==-3:
                        n3=n3+1
                    elif datain.loc[pos,'Octant']==4:
                        p4=p4+1
                    elif datain.loc[pos,'Octant']==-4:
                        n4=n4+1 
                        #updating each variables counting octants and storing in the range of mod valued difference class
                j=j+1 #loop updating

            op1=op1+p1
            on1=on1+n1
            op2=op2+p2
            on2=on2+n2
            op3=op3+p3
            on3=on3+n3
            op4=op4+p4
            on4=on4+n4
            #updating overall count of each octant value in our input datawhere it lies

            #//when loop is broken for mod range value the value is stored in the desired location using following command
            datain.loc[i+1,'1']=p1
            datain.loc[i+1,'-1']=n1
            datain.loc[i+1,'2']=p2
            datain.loc[i+1,'-2']=n2
            datain.loc[i+1,'3']=p3
            datain.loc[i+1,'-3']=n3
            datain.loc[i+1,'4']=p4
            datain.loc[i+1,'-4']=n4
            i=i+1 
            #updating another parent loop having limit noi---no. of iterations reqd
            #after parent loop also broken 
            #we are assigning the values of overall count for each octant in the whole data using following comment

        datain.loc[0,'1']=op1
        datain.loc[0,'-1']=on1
        datain.loc[0,'2']=op2
        datain.loc[0,'-2']=on2
        datain.loc[0,'3']=op3
        datain.loc[0,'-3']=on3
        datain.loc[0,'4']=op4
        datain.loc[0,'-4']=on4
        #all work done now we have desired dataframe which we need to convert in csv file using the command below
        #now finding rank based on octant appearance
        datain['Rank Octant 1']=np.nan
        datain['Rank Octant -1']=np.nan
        datain['Rank Octant 2']=np.nan
        datain['Rank Octant -2']=np.nan
        datain['Rank Octant 3']=np.nan
        datain['Rank Octant -3']=np.nan
        datain['Rank Octant 4']=np.nan
        datain['Rank Octant -4']=np.nan
        datain['Rank 1 Octant Id']=np.nan
        datain['Rank 1 Octant Name']=np.nan
        #created column for writing rank values
        datain.loc[noi+2,'Rank Octant 4']="Octant ID"
        datain.loc[noi+2,'Rank Octant -4']="Octant Name"
        datain.loc[noi+2,'Rank 1 Octant Id']="Count of Rank 1 Mod values"
        ##created required matrix 
        #now, writing program for finding rank of matrix....and filling matrix
        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
            #this dictionary has been made to name the corresponding octant
        datain.loc[noi+3,'Rank Octant 4']="1"
        datain.loc[noi+4,'Rank Octant 4']="-1"
        datain.loc[noi+5,'Rank Octant 4']="2"
        datain.loc[noi+6,'Rank Octant 4']="-2"
        datain.loc[noi+7,'Rank Octant 4']="3"
        datain.loc[noi+8,'Rank Octant 4']="-3"
        datain.loc[noi+9,'Rank Octant 4']="4"
        datain.loc[noi+10,'Rank Octant 4']="-4"
        datain.loc[noi+3,'Rank Octant -4']=octant_name_id_mapping["1"]
        datain.loc[noi+4,'Rank Octant -4']=octant_name_id_mapping["-1"]
        datain.loc[noi+5,'Rank Octant -4']=octant_name_id_mapping["2"]
        datain.loc[noi+6,'Rank Octant -4']=octant_name_id_mapping["-2"]
        datain.loc[noi+7,'Rank Octant -4']=octant_name_id_mapping["3"]
        datain.loc[noi+8,'Rank Octant -4']=octant_name_id_mapping["-3"]
        datain.loc[noi+9,'Rank Octant -4']=octant_name_id_mapping["4"]
        datain.loc[noi+10,'Rank Octant -4']=octant_name_id_mapping["-4"]
        #filled matrix required as according to tutorial corresponding to their requirement #no logic used used till now
        i=0 #we will do iteration for the all rank count and write the count for one iteration...
        r1cp1=r1cn1=r1cp2=r1cn2=r1cp3=r1cn3=r1cp4=r1cn4=0 
        #variable initiated for rank 1 count p(positive)/n(negative) octant id we will use it later for rank 1 appearance in the mod values of diff. octant id
        while (i<noi+1):
            rank_dict = {datain.loc[i,'1'] :"1",datain.loc[i,'-1'] :"-1",datain.loc[i,'2'] :"2",datain.loc[i,'-2'] :"-2",datain.loc[i,'3'] :"3",datain.loc[i,'-3'] :"-3",datain.loc[i,'4'] :"4",datain.loc[i,'-4'] :"-4"}
            #initiated dictionary of values at corresponding row iteration at different columns with their column name as value
            rank_list =[datain.loc[i,'1'],datain.loc[i,'-1'],datain.loc[i,'2'],datain.loc[i,'-2'],datain.loc[i,'3'],datain.loc[i,'-3'],datain.loc[i,'4'],datain.loc[i,'-4'] ]
            #created list and sorted in next step to get the rank on the basis of indexes easily....
            rank_list.sort() 
            for j in range(8):
                lok=rank_dict[rank_list[j]] #made variable to use it frequently with ease
                if (lok=="1"):
                    datain.loc[i,'Rank Octant 1']=8-j
                if (lok=="-1"):
                    datain.loc[i,'Rank Octant -1']=8-j
                if (lok=="2"):
                    datain.loc[i,'Rank Octant 2']=8-j
                if (lok=="-2"):
                    datain.loc[i,'Rank Octant -2']=8-j
                if (lok=="3"):
                    datain.loc[i,'Rank Octant 3']=8-j
                if (lok=="-3"):
                    datain.loc[i,'Rank Octant -3']=8-j
                if (lok=="4"):
                    datain.loc[i,'Rank Octant 4']=8-j
                if (lok=="-4"):
                    datain.loc[i,'Rank Octant -4']=8-j
                    #assigning rank values of each octant in each mod value range as well as overall data in its corresponding column
            datain.loc[i,'Rank 1 Octant Id']=rank_dict[rank_list[7]]
            #since out list are sorted in ascending order so we are traversing from back side to get the max value in the list to assign it as rank 1
            datain.loc[i,'Rank 1 Octant Name']=octant_name_id_mapping[datain.loc[i,'Rank 1 Octant Id']]
            ##now we will be counting rant one appearance of each octant...
            if(i>0):
                #this we have done beacause we have to count the rank one value for each mod duration only
                if(datain.loc[i,'Rank 1 Octant Id']=="1"):
                    r1cp1+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="-1"):
                    r1cn1+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="2"):
                    r1cp2+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="-2"):
                    r1cn2+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="3"):
                    r1cp3+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="-3"):
                    r1cn3+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="4"):
                    r1cp4+=1
                if(datain.loc[i,'Rank 1 Octant Id']=="-4"):
                    r1cn4+=1
                    #all values will be updated at end of the iteration
            i=i+1
        datain.loc[noi+3,'Rank 1 Octant Id']=r1cp1
        datain.loc[noi+4,'Rank 1 Octant Id']=r1cn1
        datain.loc[noi+5,'Rank 1 Octant Id']=r1cp2
        datain.loc[noi+6,'Rank 1 Octant Id']=r1cn2
        datain.loc[noi+7,'Rank 1 Octant Id']=r1cp3
        datain.loc[noi+8,'Rank 1 Octant Id']=r1cn3
        datain.loc[noi+9,'Rank 1 Octant Id']=r1cp4
        datain.loc[noi+10,'Rank 1 Octant Id']=r1cn4
        #assigning value in the required matrix with the help of variable initiated earlier

        ##MAKING OTHER REQUIRED MATRIX..........
        #transition count..
        datain['']=np.nan
        datain['remove0']=np.nan
        datain['remove1']=np.nan
        datain['remove2']=np.nan
        datain['remove3']=np.nan
        datain['remove4']=np.nan
        datain['remove5']=np.nan
        datain['remove6']=np.nan
        datain['remove7']=np.nan
        datain['remove8']=np.nan
        datain['remove9']=np.nan

        #making matrix
        datain.loc[1,'remove0']="From"
        datain.loc[0,'remove1']="Octant #"
        datain.loc[1,'remove1']='+1'
        datain.loc[2,'remove1']='-1'
        datain.loc[3,'remove1']='+2'
        datain.loc[4,'remove1']='-2'
        datain.loc[5,'remove1']='+3'
        datain.loc[6,'remove1']='-3'
        datain.loc[7,'remove1']='+4'
        datain.loc[8,'remove1']='-4'
        datain.loc[0,'remove2']='+1'
        datain.loc[0,'remove3']='-1'
        datain.loc[0,'remove4']='+2'
        datain.loc[0,'remove5']='-2'
        datain.loc[0,'remove6']='+3'
        datain.loc[0,'remove7']='-3'
        datain.loc[0,'remove8']='+4'
        datain.loc[0,'remove9']='-4'
        #overall transition matrix made.....
        #now working for assigning value in it....
        #filling 0 in overall transition matrix
        for i in range(1,9):
            datain.loc[i,'remove2']=0
            datain.loc[i,'remove3']=0
            datain.loc[i,'remove4']=0
            datain.loc[i,'remove5']=0
            datain.loc[i,'remove6']=0
            datain.loc[i,'remove7']=0
            datain.loc[i,'remove8']=0
            datain.loc[i,'remove9']=0
        pos=12 #storing for position in a variable later to use in loop
        #for mod transition
        i=0
        while i<noi: 
            ll=i*mod
            if ((i+1)*mod)>=total_size:
                ul=total_size-1
            else:
                ul=(i+1)*mod-1
                #defined upper limit as ul and lower limit as ll
            #making mod transition matrix
            datain.loc[pos,'remove1']="Mod Transition Count"
            datain.loc[pos+1,'remove1']="{}-{}".format(ll,ul)
            datain.loc[pos+1,'remove2']='To'
            datain.loc[pos+2,'remove1']="Octant #"
            datain.loc[pos+2,'remove2']='+1'
            datain.loc[pos+2,'remove3']='-1'
            datain.loc[pos+2,'remove4']='+2'
            datain.loc[pos+2,'remove5']='-2'
            datain.loc[pos+2,'remove6']='+3'
            datain.loc[pos+2,'remove7']='-3'
            datain.loc[pos+2,'remove8']='+4'
            datain.loc[pos+2,'remove9']='-4'
            datain.loc[pos+3,'remove0']="From"
            datain.loc[pos+3,'remove1']="+1"
            datain.loc[pos+4,'remove1']="-1"
            datain.loc[pos+5,'remove1']="+2"
            datain.loc[pos+6,'remove1']="-2"
            datain.loc[pos+7,'remove1']="+3"
            datain.loc[pos+8,'remove1']="-3"
            datain.loc[pos+9,'remove1']="+4"
            datain.loc[pos+10,'remove1']="-4"
            #matrix prepared
            #assigning 0 to the each mod matrix
            for j in range(3+pos,11+pos):
                datain.loc[j,'remove2']=0
                datain.loc[j,'remove3']=0
                datain.loc[j,'remove4']=0
                datain.loc[j,'remove5']=0
                datain.loc[j,'remove6']=0
                datain.loc[j,'remove7']=0
                datain.loc[j,'remove8']=0
                datain.loc[j,'remove9']=0
            for z in range (ll,ul+1):
                
                if(datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+3,'remove2']+=1 #for mod transition
                    datain.loc[1,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+3,'remove3']+=1
                    datain.loc[1,'remove3']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+3,'remove4']+=1
                    datain.loc[1,'remove4']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+3,'remove5']+=1
                    datain.loc[1,'remove5']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+3,'remove6']+=1
                    datain.loc[1,'remove6']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+3,'remove7']+=1
                    datain.loc[1,'remove7']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+3,'remove8']+=1
                    datain.loc[1,'remove8']+=1
                elif (datain.loc[z,'Octant']==1 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+3,'remove9']+=1
                    datain.loc[1,'remove9']+=1
                #updated row corresponding to from +1 to any transition (mod or overall) 
                #now for -1
                if(datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+4,'remove2']+=1 #for mod transition
                    datain.loc[2,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+4,'remove3']+=1
                    datain.loc[2,'remove3']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+4,'remove4']+=1
                    datain.loc[2,'remove4']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+4,'remove5']+=1
                    datain.loc[2,'remove5']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+4,'remove6']+=1
                    datain.loc[2,'remove6']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+4,'remove7']+=1
                    datain.loc[2,'remove7']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+4,'remove8']+=1
                    datain.loc[2,'remove8']+=1
                elif (datain.loc[z,'Octant']==-1 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+4,'remove9']+=1
                    datain.loc[2,'remove9']+=1
                #updated row corresponding to from -1 to any transition (mod or overall) 
                #now for +2
                if(datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+5,'remove2']+=1 #for mod transition
                    datain.loc[3,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+5,'remove3']+=1
                    datain.loc[3,'remove3']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+5,'remove4']+=1
                    datain.loc[3,'remove4']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+5,'remove5']+=1
                    datain.loc[3,'remove5']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+5,'remove6']+=1
                    datain.loc[3,'remove6']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+5,'remove7']+=1
                    datain.loc[3,'remove7']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+5,'remove8']+=1
                    datain.loc[3,'remove8']+=1
                elif (datain.loc[z,'Octant']==2 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+5,'remove9']+=1
                    datain.loc[3,'remove9']+=1
                #updated row corresponding to from +2 to any transition (mod or overall)  

                #now for -2
                if(datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+6,'remove2']+=1 #for mod transition
                    datain.loc[4,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+6,'remove3']+=1
                    datain.loc[4,'remove3']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+6,'remove4']+=1
                    datain.loc[4,'remove4']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+6,'remove5']+=1
                    datain.loc[4,'remove5']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+6,'remove6']+=1
                    datain.loc[4,'remove6']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+6,'remove7']+=1
                    datain.loc[4,'remove7']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+6,'remove8']+=1
                    datain.loc[4,'remove8']+=1
                elif (datain.loc[z,'Octant']==-2 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+6,'remove9']+=1
                    datain.loc[4,'remove9']+=1
                #updated row corresponding to from -2 to any transition (mod or overall)

                #now for +3
                if(datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+7,'remove2']+=1 #for mod transition
                    datain.loc[5,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+7,'remove3']+=1
                    datain.loc[5,'remove3']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+7,'remove4']+=1
                    datain.loc[5,'remove4']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+7,'remove5']+=1
                    datain.loc[5,'remove5']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+7,'remove6']+=1
                    datain.loc[5,'remove6']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+7,'remove7']+=1
                    datain.loc[5,'remove7']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+7,'remove8']+=1
                    datain.loc[5,'remove8']+=1
                elif (datain.loc[z,'Octant']==3 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+7,'remove9']+=1
                    datain.loc[5,'remove9']+=1
                #updated row corresponding to from +3 to any transition (mod or overall) 

                #now for -3
                if(datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+8,'remove2']+=1 #for mod transition
                    datain.loc[6,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+8,'remove3']+=1
                    datain.loc[6,'remove3']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+8,'remove4']+=1
                    datain.loc[6,'remove4']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+8,'remove5']+=1
                    datain.loc[6,'remove5']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+8,'remove6']+=1
                    datain.loc[6,'remove6']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+8,'remove7']+=1
                    datain.loc[6,'remove7']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+8,'remove8']+=1
                    datain.loc[6,'remove8']+=1
                elif (datain.loc[z,'Octant']==-3 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+8,'remove9']+=1
                    datain.loc[6,'remove9']+=1
                #updated row corresponding to from -3 to any transition (mod or overall)     
                #now for +4
                if(datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+9,'remove2']+=1 #for mod transition
                    datain.loc[7,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+9,'remove3']+=1
                    datain.loc[7,'remove3']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+9,'remove4']+=1
                    datain.loc[7,'remove4']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+9,'remove5']+=1
                    datain.loc[7,'remove5']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+9,'remove6']+=1
                    datain.loc[7,'remove6']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+9,'remove7']+=1
                    datain.loc[7,'remove7']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+9,'remove8']+=1
                    datain.loc[7,'remove8']+=1
                elif (datain.loc[z,'Octant']==4 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+9,'remove9']+=1
                    datain.loc[7,'remove9']+=1
                #updated row corresponding to from +4 to any transition (mod or overall)  
                #now for -4
                if(datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==1):
                    datain.loc[pos+10,'remove2']+=1 #for mod transition
                    datain.loc[8,'remove2']+=1 #for overall transition
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==-1):
                    datain.loc[pos+10,'remove3']+=1
                    datain.loc[8,'remove3']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==2):
                    datain.loc[pos+10,'remove4']+=1
                    datain.loc[8,'remove4']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==-2):
                    datain.loc[pos+10,'remove5']+=1
                    datain.loc[8,'remove5']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==3):
                    datain.loc[pos+10,'remove6']+=1
                    datain.loc[8,'remove6']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==-3):
                    datain.loc[pos+10,'remove7']+=1
                    datain.loc[8,'remove7']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==4):
                    datain.loc[pos+10,'remove8']+=1
                    datain.loc[8,'remove8']+=1
                elif (datain.loc[z,'Octant']==-4 and datain.loc[z+1,'Octant']==-4):
                    datain.loc[pos+10,'remove9']+=1
                    datain.loc[8,'remove9']+=1
                    #updated row corresponding to from -4 to any transition (mod or overall)  
                if(z==total_size-2):
                    break   
                    #here we used break statement because we need to traverse to total no. of rows minus 1 because indexing in python is from zero
                    #and due to this last element of row will be at index total size -1
                    #and we need to traverse to 2nd last row only so breaking loop at 2nd last element
                
            pos=pos+13 #updating position for next iteration
            i=i+1
        ####################################################
        #now doing longest susequent count part...........
        datain['']=np.nan
        datain['Octant ##']=np.nan
        datain['Longest Subsequence Length']=np.nan
        datain['Count']=np.nan
        datain.loc[0,'Octant ##']='+1'
        datain.loc[1,'Octant ##']='-1'
        datain.loc[2,'Octant ##']='+2'
        datain.loc[3,'Octant ##']='-2'
        datain.loc[4,'Octant ##']='+3'
        datain.loc[5,'Octant ##']='-3'
        datain.loc[6,'Octant ##']='+4'
        datain.loc[7,'Octant ##']='-4'
        #till here matrix has been made now we have to fill the values in the dataframe
        scp1=scn1=scp2=scn2=scp3=scn3=scp4=scn4=0 #made variable for subsequent positive and negative count
        cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0 #made variables for count the appearance of subsequent count
        listp1=[]
        listn1=[]
        listp2=[]
        listn2=[]
        listp3=[]
        listn3=[]
        listp4=[]
        listn4=[]
        #empty list made to use it later for time range printing

        for i in range(0,total_size):
            if(datain.loc[i,'Octant']==1): 
                count=0 #initiated local variable which will help in updating sc#@(p/n)(1/2/3/4)
                while (datain.loc[i,'Octant']==1): #loop for checking appearance of velocity in octant
                    count+=1 #updating count
                    i+=1 #updating rows for checking continuity in a octant
                    if(i>=total_size):
                        break #this is because we have only rows upto index total size-1 and .loc func will unable to access the rows which actually does not exist
                if(count>scp1):
                    scp1=count #updating subsequent longest  count 
                    cp1=1 #assigning value as 1 because of new longest subsequent count
                    listp1.clear() #list has been cleared in case it has any previous element which shouldn't in case of if case
                    listp1.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scp1):
                    cp1+=1 #updating count value of subsequent longest count in case of same subs. long. count
                    listp1.append(i-1) #storing index of upper bound(inclusive) which will help later
                    ################
                    #after this same code has been done for all the octanct type occurence so no comment needed till code line 166
            elif(datain.loc[i,'Octant']==-1):
                count=0
                while (datain.loc[i,'Octant']==-1):
                    count+=1 #explained in +ve 1 coode writing all step is same as that
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scn1):
                    scn1=count
                    cn1=1
                    listn1.clear()
                    listn1.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scn1):
                    cn1+=1
                    listn1.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==2):
                count=0
                while (datain.loc[i,'Octant']==2):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scp2):
                    scp2=count
                    cp2=1
                    listp2.clear()
                    listp2.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scp2):
                    cp2+=1
                    listp2.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==-2):
                count=0
                while (datain.loc[i,'Octant']==-2):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scn2):
                    scn2=count
                    cn2=1
                    listn2.clear()
                    listn2.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scn2):
                    cn2+=1
                    listn2.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==3):
                count=0
                while (datain.loc[i,'Octant']==3):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scp3):
                    scp3=count
                    cp3=1
                    listp3.clear()
                    listp3.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scp3):
                    cp3+=1
                    listp3.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==-3):
                count=0
                while (datain.loc[i,'Octant']==-3):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scn3):
                    scn3=count
                    cn3=1
                    listn3.clear()
                    listn3.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scn3):
                    cn3+=1
                    listn3.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==4):
                count=0
                while (datain.loc[i,'Octant']==4):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scp4):
                    scp4=count
                    cp4=1
                    listp4.clear()
                    listp4.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scp4):
                    cp4+=1
                    listp4.append(i-1) #storing index of upper bound(inclusive) which will help later
            elif(datain.loc[i,'Octant']==-4):
                count=0
                while (datain.loc[i,'Octant']==-4):
                    count+=1
                    i+=1
                    if(i>=total_size):
                        break
                if(count>scn4):
                    scn4=count
                    cn4=1
                    listn4.clear()
                    listn4.append(i-1) #storing index of upper bound(inclusive) which will help later
                elif (count==scn4):
                    cn4+=1
                    listn4.append(i-1) #storing index of upper bound(inclusive) which will help later
            i=i-1 #this substraction in the row has been done because of an extra addition of row while checking the repitition
            #in the octant and in case of failing of repitition condition row will gone to another position already
            #we know about loop condition after each iteration it increases its loop by doing one iteration operation
            #so on adding 1 we will actually jump two row but we need to jump only one row
            #1 has been substracted to have no skipped row to get correct result
            ###till here done#######
        datain.loc[0,'Longest Subsequence Length']=scp1 #assigning longest subsequent value corresponding to positive 1 octant value
        datain.loc[0,'Count']=cp1 #assigning count of longest subs. value corresponding to +ve 1
        datain.loc[1,'Longest Subsequence Length']=scn1
        datain.loc[1,'Count']=cn1 #assigning count of longest subs. value corresponding to -ve 1
        datain.loc[2,'Longest Subsequence Length']=scp2
        datain.loc[2,'Count']=cp2 #assigning count of longest subs. value corresponding to +ve 2
        datain.loc[3,'Longest Subsequence Length']=scn2
        datain.loc[3,'Count']=cn2 #assigning count of longest subs. value corresponding to -ve 2
        datain.loc[4,'Longest Subsequence Length']=scp3
        datain.loc[4,'Count']=cp3 #assigning count of longest subs. value corresponding to +ve 3
        datain.loc[5,'Longest Subsequence Length']=scn3
        datain.loc[5,'Count']=cn3 #assigning count of longest subs. value corresponding to -ve 3
        datain.loc[6,'Longest Subsequence Length']=scp4
        datain.loc[6,'Count']=cp4 #assigning count of longest subs. value corresponding to +ve 4
        datain.loc[7,'Longest Subsequence Length']=scn4
        datain.loc[7,'Count']=cn4 #assigning count of longest subs. value corresponding to -ve 4


        #till here done now working  for range values and indexe range of longest subsequence count
        datain['']=np.nan #created blank column with no name as asked in the excel file
        datain['Octant ###']=np.nan
        datain['Longest Subsquence Length#']=np.nan
        datain['Count#']=np.nan
        #making desired matrix for +1
        datain.loc[0,'Octant ###']='+1'
        datain.loc[1,'Octant ###']='Time'
        datain.loc[0,'Longest Subsquence Length#']=scp1
        datain.loc[1,'Longest Subsquence Length#']='From'
        datain.loc[0,'Count#']=cp1
        datain.loc[1,'Count#']='To'
        #making desired matrix for filling the time range of -1
        datain.loc[2+cp1,'Octant ###']='-1'
        datain.loc[2+cp1,'Longest Subsquence Length#']=scn1
        datain.loc[2+cp1,'Count#']=cn1
        datain.loc[3+cp1,'Octant ###']='Time'
        datain.loc[3+cp1,'Longest Subsquence Length#']='From'
        datain.loc[3+cp1,'Count#']='To'
        #making desired matrix for filling the time range of +2
        datain.loc[4+cp1+cn1,'Octant ###']='+2'
        datain.loc[4+cp1+cn1,'Longest Subsquence Length#']=scp2
        datain.loc[4+cp1+cn1,'Count#']=cp2
        datain.loc[5+cp1+cn1,'Octant ###']='Time'
        datain.loc[5+cp1+cn1,'Longest Subsquence Length#']='From'
        datain.loc[5+cp1+cn1,'Count#']='To'
        #making desired matrix for filling the time range of -2
        datain.loc[6+cp1+cn1+cp2,'Octant ###']='-2'
        datain.loc[6+cp1+cn1+cp2,'Longest Subsquence Length#']=scn2
        datain.loc[6+cp1+cn1+cp2,'Count#']=cn2
        datain.loc[7+cp1+cn1+cp2,'Octant ###']='Time'
        datain.loc[7+cp1+cn1+cp2,'Longest Subsquence Length#']='From'
        datain.loc[7+cp1+cn1+cp2,'Count#']='To'
        #making desired matrix for filling the time range of +3
        datain.loc[8+cp1+cn1+cp2+cn2,'Octant ###']='+3'
        datain.loc[8+cp1+cn1+cp2+cn2,'Longest Subsquence Length#']=scp3
        datain.loc[8+cp1+cn1+cp2+cn2,'Count#']=cp3
        datain.loc[9+cp1+cn1+cp2+cn2,'Octant ###']='Time'
        datain.loc[9+cp1+cn1+cp2+cn2,'Longest Subsquence Length#']='From'
        datain.loc[9+cp1+cn1+cp2+cn2,'Count#']='To'
        #making desired matrix for filling the time range of -3
        datain.loc[10+cp1+cn1+cp2+cn2+cp3,'Octant ###']='-3'
        datain.loc[10+cp1+cn1+cp2+cn2+cp3,'Longest Subsquence Length#']=scn3
        datain.loc[10+cp1+cn1+cp2+cn2+cp3,'Count#']=cn3
        datain.loc[11+cp1+cn1+cp2+cn2+cp3,'Octant ###']='Time'
        datain.loc[11+cp1+cn1+cp2+cn2+cp3,'Longest Subsquence Length#']='From'
        datain.loc[11+cp1+cn1+cp2+cn2+cp3,'Count#']='To'
        #making desired matrix for filling the time range of +4
        datain.loc[12+cp1+cn1+cp2+cn2+cp3+cn3,'Octant ###']='+4'
        datain.loc[12+cp1+cn1+cp2+cn2+cp3+cn3,'Longest Subsquence Length#']=scp4
        datain.loc[12+cp1+cn1+cp2+cn2+cp3+cn3,'Count#']=cp4
        datain.loc[13+cp1+cn1+cp2+cn2+cp3+cn3,'Octant ###']='Time'
        datain.loc[13+cp1+cn1+cp2+cn2+cp3+cn3,'Longest Subsquence Length#']='From'
        datain.loc[13+cp1+cn1+cp2+cn2+cp3+cn3,'Count#']='To'
        #making desired matrix for filling the time range of -4
        datain.loc[14+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Octant ###']='-4'
        datain.loc[14+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Longest Subsquence Length#']=scn4
        datain.loc[14+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Count#']=cn4
        datain.loc[15+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Octant ###']='Time'
        datain.loc[15+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Longest Subsquence Length#']='From'
        datain.loc[15+cp1+cn1+cp2+cn2+cp3+cn3+cp4,'Count#']='To'

        ##############################################################################
        #till here matrix done
        #now working for matrix filling
        for c in range (0,cp1):
            datain.loc[2+c,'Count#']=datain.at[listp1[c],'T'] #upper limit time i.e. to
            datain.loc[2+c,'Longest Subsquence Length#']=datain.at[1+listp1[c]-scp1,'T'] #lower limit time i. e. from
        for c in range (0,cn1):
            datain.loc[4+cp1+c,'Count#']=datain.at[listn1[c],'T'] #upper limit time i.e. to
            datain.loc[4+cp1+c,'Longest Subsquence Length#']=datain.at[1+listn1[c]-scn1,'T'] #lower limit time i. e. from
        for c in range (0,cp2):
            datain.loc[6+cp1+cn1+c,'Count#']=datain.at[listp2[c],'T'] #upper limit time i.e. to
            datain.loc[6+cp1+cn1+c,'Longest Subsquence Length#']=datain.at[1+listp2[c]-scp2,'T'] #lower limit time i. e. from
        for c in range (0,cn2):
            datain.loc[8+cp1+cn1+cp2+c,'Count#']=datain.at[listn2[c],'T'] #upper limit time i.e. to
            datain.loc[8+cp1+cn1+cp2+c,'Longest Subsquence Length#']=datain.at[1+listn2[c]-scn2,'T'] #lower limit time i. e. from
        for c in range (0,cp3):
            datain.loc[10+cp1+cn1+cp2+cn2+c,'Count#']=datain.at[listp3[c],'T'] #upper limit time i.e. to
            datain.loc[10+cp1+cn1+cp2+cn2+c,'Longest Subsquence Length#']=datain.at[1+listp3[c]-scp3,'T'] #lower limit time i. e. from
        for c in range (0,cn3):
            datain.loc[12+cp1+cn1+cp2+cn2+cp3+c,'Count#']=datain.at[listn3[c],'T'] #upper limit time i.e. to
            datain.loc[12+cp1+cn1+cp2+cn2+cp3+c,'Longest Subsquence Length#']=datain.at[1+listn3[c]-scn3,'T'] #lower limit time i. e. from
        for c in range (0,cp4):
            datain.loc[14+cp1+cn1+cp2+cn2+cp3+cn3+c,'Count#']=datain.at[listp4[c],'T'] #upper limit time i.e. to
            datain.loc[14+cp1+cn1+cp2+cn2+cp3+cn3+c,'Longest Subsquence Length#']=datain.at[1+listp4[c]-scp4,'T'] #lower limit time i. e. from
        for c in range (0,cn4):
            datain.loc[16+cp1+cn1+cp2+cn2+cp3+cn3+cp4+c,'Count#']=datain.at[listn4[c],'T'] #upper limit time i.e. to
            datain.loc[16+cp1+cn1+cp2+cn2+cp3+cn3+cp4+c,'Longest Subsquence Length#']=datain.at[1+listn4[c]-scn4,'T'] #lower limit time i. e. from
        
        # Insering empty columns at desired positions
        datain.insert(32,"",np.nan,True)
        datain.insert(43,"",np.nan,True)
        datain.insert(47,"",np.nan,True)

        # Renaming columnss
        datain.rename(columns={'remove_mod':''}, inplace=True)
        datain.rename(columns={'remove0':''}, inplace=True)
        datain.rename(columns={'remove1':'Overall transition Count'}, inplace=True)
        datain.rename(columns={'remove2':''}, inplace=True)
        datain.rename(columns={'remove3':''}, inplace=True)
        datain.rename(columns={'remove4':''}, inplace=True)
        datain.rename(columns={'remove5':''}, inplace=True)
        datain.rename(columns={'remove6':''}, inplace=True)
        datain.rename(columns={'remove7':''}, inplace=True)
        datain.rename(columns={'remove8':''}, inplace=True)
        datain.rename(columns={'remove9':''}, inplace=True)

        
        file_name = os.path.basename(file)

        # file name without extension
        name = os.path.splitext(file_name)[0]
        # Saving file
        datain.to_excel('output\\'+name+ f'_octant_analysis_mod_{mod}.xlsx',index=False) # it makes a csv file with the name given in 'quote'
        #index = false do not make make columns for index values as we have no requirement of it in this case
        # total_count = datain['Longest Subsquence Length#'].size

        # Imporing libraries
        from openpyxl.styles import Border,Side
        # load workbook
        workbook= openpyxl.load_workbook('output\\'+name+ f'_octant_analysis_mod_{mod}.xlsx')
        worksheet = workbook['Sheet1']
        # Creating border style
        top = Side(border_style='thin', color="000000")
        bottom = Side(border_style='thin', color="000000")
        right = Side(border_style='thin', color="000000")
        left = Side(border_style='thin', color="000000")
        border = Border(top=top, bottom = bottom, right=right, left=left)

        # Applying border to ranges
        grid1=worksheet['N1':f'AF{noi+2}']
        for cell in grid1:
            for x in cell:
                x.border=border
        # Applying border to ranges
        grid1 = worksheet[f'AC{noi+4}':f'AE{noi+12}']
        for cell in grid1:
            for x in cell:
                x.border=border

        # Applying border to ranges
        grid1 = worksheet[f'AI2':f'AQ10']
        for cell in grid1:
            for x in cell:
                x.border=border
        ex = 0
        # Applying border to ranges
        while ex<noi:
            grid1 = worksheet[f'AI{16+13*ex}':f'AQ{24+13*ex}']
            for cell in grid1:
                for x in cell:
                    x.border=border
            ex =ex+1
        # Applying border to ranges
        grid1 = worksheet[f'AS1':f'AU9']
        for cell in grid1:
            for x in cell:
                x.border=border
        # Applying border to ranges
        grid1 = worksheet[f'AW1':f'AY{17+cp1+cn1+cp2+cn2+cp3+cn3+cp4+cn4}']
        for cell in grid1:
            for x in cell:
                x.border=border
        # Applying background colors 
        fillcolur = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font = Font(bold=False, color='000000')
        dxf=DifferentialStyle(font=font ,fill=fillcolur)
        rule = Rule(type='cellIs', operator='equal',formula=[1,1], dxf=dxf)
        # Highlighting cell ranges
        for i in range(8):
            grid_cell = f'$AJ${3+i}:$AQ${3+i}'
            rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+grid_cell+")"])
            worksheet.conditional_formatting.add(grid_cell, rule1)
        # Highlighting cell ranges
        for q in range(noi):
            for i in range(8):
                grid_cell = f'$AJ${17+13*q+i}:$AQ${17+13*q+i}'
                rule1 = Rule(type='cellIs',operator='equal', dxf=dxf, formula=["=MAX("+grid_cell+")"])
                worksheet.conditional_formatting.add(grid_cell, rule1)
        
        # Adding highlights
        for i in range(noi+1):
            worksheet.conditional_formatting.add(f'W{2+i}:AD{2+i}', rule)
        # Saving workbook
        workbook.save('output\\'+name+ f'_octant_analysis_mod_{mod}.xlsx')
    except FileNotFoundError:
        print("Hey...file inputed by user is not found")


	

##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000

path = r'C:\Users\pc\Documents\GitHub\2001CE45_2022\tut07'
dir_list = os.listdir(path+'\input')
for file in dir_list:
    octant_analysis( file, mod)
    # file name with extension
    





#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
